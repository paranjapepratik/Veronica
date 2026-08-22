"""
Veronica — AI-Powered Literature Review Tool  v4.0
Sources : PubMed (Entrez) + Europe PMC + OpenAlex — all keyless
AI      : Ollama, local

What changed from v3 (see CHANGELOG.md):
  * retrieval budget is real (v3 passed show_n where fetch_n belonged)
  * concept-based query builder with MeSH + [tiab] OR-groups, query shown in UI
  * papers without full text are KEPT and flagged, not silently dropped
  * local BM25 pre-rank, then LLM scoring of the top slice only
  * rubric scoring: population / intervention / outcome / design, 0-3 each,
    each with a quoted evidence span. Total 0-12, auditable.
  * concurrent Ollama calls with keep_alive
  * theme clustering, include/exclude screening, drafted synthesis
  * light + dark UI themes

Setup:
    pip3 install requests openpyxl biopython
    curl -fsSL https://ollama.com/install.sh | sh
    ollama pull llama3.1:8b

Run:
    python3 veronica.py
"""

import os, re, sys, json, math, time, queue, zipfile, threading, datetime, webbrowser
import tkinter as tk
from tkinter import ttk, messagebox
from pathlib import Path
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from Bio import Entrez, Medline

VERSION       = "5.0"
DESKTOP       = Path.home() / "Desktop"
BASE_DIR      = DESKTOP / "Veronica"
TODAY         = datetime.date.today().strftime("%Y-%m-%d")
OLLAMA_HOST   = "http://localhost:11434"
DEFAULT_MODEL = "llama3.1:8b"
SCORE_MODEL   = "llama3.2:3b"     # scoring is a triage job: small and fast beats big
UA            = {"User-Agent": f"Veronica/{VERSION} (literature review tool; mailto:research@local)"}
Entrez.email  = "veronica.tool@research.local"
SETTINGS_PATH = Path.home() / ".veronica.json"

# ══════════════════════════════════════════════════════════════════════════════
#  Theme
# ══════════════════════════════════════════════════════════════════════════════
THEMES = {
    "light": dict(
        bg="#f2f2f3", surface="#e9e9ea", surface2="#dedee0",
        line="#c9c9cc", line2="#e0e0e2",
        text="#1d1f20", mut="#5d5d60", mut2="#7a7a7d",
        accent="#5980a6", accent_dim="#93aec9", on_accent="#f2f2f3",
        sel="#d6e2ee", quote="#424244", bar_off="#d4d4d7",
        ok="#3f7d4f", warn="#a6743f", err="#9d4b4b",
        head="Barlow Condensed", body="Barlow", mono="Menlo",
    ),
    "dark": dict(
        bg="#0f0f0f", surface="#181818", surface2="#222222",
        line="#2a2a2a", line2="#1c1c1c",
        text="#e8e4dc", mut="#999990", mut2="#555550",
        accent="#c4a96a", accent_dim="#7a6540", on_accent="#0e0e0e",
        sel="#1f1c12", quote="#999990", bar_off="#2a2a2a",
        ok="#5aaa5a", warn="#c4a96a", err="#c46a6a",
        head="Georgia", body="Georgia", mono="Menlo",
    ),
}

def font_or(family, fallbacks, size, weight="normal"):
    """tkinter silently falls back on missing families; keep a sane order."""
    try:
        from tkinter import font as tkfont
        available = set(f.lower() for f in tkfont.families())
        for f in [family] + fallbacks:
            if f.lower() in available:
                return (f, size, weight)
    except Exception:
        pass
    return (fallbacks[-1] if fallbacks else family, size, weight)


# ══════════════════════════════════════════════════════════════════════════════
#  Query building
# ══════════════════════════════════════════════════════════════════════════════
SYNONYMS = {
    "alzheimer disease": ["alzheimer*", "AD"],
    "alzheimer's disease": ["alzheimer*", "AD"],
    "mild cognitive impairment": ["MCI", "prodromal dementia", "cognitive dysfunction"],
    "mouse model": ["mouse model*", "murine model*", "transgenic mice", "mice"],
    "randomized controlled trial": ["RCT", "randomised controlled trial"],
    "machine learning": ["deep learning", "neural network*"],
    "covid-19": ["SARS-CoV-2", "coronavirus disease 2019"],
    "type 2 diabetes": ["T2DM", "non-insulin-dependent diabetes"],
    "gut microbiome": ["gut microbiota", "intestinal microbiome"],
    "large language model": ["LLM", "language model", "transformer"],
    "neural network": ["deep neural network", "DNN", "deep learning"],
    "reinforcement learning": ["RL", "policy learning"],
    "lithium ion battery": ["Li-ion battery", "lithium-ion cell"],
    "solar cell": ["photovoltaic", "PV cell"],
    "quantum computing": ["quantum computer", "qubit"],
    "climate change": ["global warming", "climate variability"],
    "higher education": ["university", "tertiary education"],
    "qualitative research": ["qualitative study", "interview study"],
}

STOP = set("""a an the and or of in on for to with without from by at as is are was were be been being
this that these those we our their its it he she they i study studies paper papers using used use
results result method methods conclusion conclusions background objective objectives purpose aim aims
significant significantly show shows showed shown found find finding findings between among within
during after before both such more most less least than then thus therefore however although while
can could may might will would also have has had do does did not no nor but if per via versus vs
patients patient group groups control controls level levels effect effects associated association
increase increased decrease decreased high low new novel""".split())


def parse_concepts(raw: str) -> list[str]:
    """One concept per line, or comma-separated. Concepts are ANDed."""
    parts = re.split(r"[\n;,]+", raw or "")
    return [p.strip() for p in parts if p.strip()]


def concept_variants(concept: str) -> list[str]:
    key = concept.strip().lower()
    out = [concept.strip()]
    for syn in SYNONYMS.get(key, []):
        if syn.lower() not in [o.lower() for o in out]:
            out.append(syn)
    return out


def build_pubmed_query(concepts: list[str], year_from: str = "", art_type: str = "") -> str:
    """Each concept becomes a MeSH + [tiab] OR-group; groups are ANDed."""
    groups = []
    for c in concepts:
        variants = concept_variants(c)
        terms = [f'"{c}"[MeSH Terms]']
        for v in variants:
            terms.append(f'{v}[tiab]' if "*" in v else f'"{v}"[tiab]')
        groups.append("(" + " OR ".join(terms) + ")")
    q = " AND ".join(groups) if groups else ""
    if art_type:
        q += f' AND "{art_type}"[Publication Type]'
    if year_from.strip().isdigit():
        q += f" AND {year_from.strip()}:3000[pdat]"
    return q


def build_epmc_query(concepts: list[str], year_from: str = "") -> str:
    groups = []
    for c in concepts:
        variants = concept_variants(c)
        terms = [f'"{v}"' if "*" not in v else v for v in variants]
        groups.append("(" + " OR ".join(terms) + ")")
    q = " AND ".join(groups) if groups else ""
    if year_from.strip().isdigit():
        q += f" AND (FIRST_PDATE:[{year_from.strip()}-01-01 TO 3000-12-31])"
    return q


# ══════════════════════════════════════════════════════════════════════════════
#  Sources
# ══════════════════════════════════════════════════════════════════════════════
def blank_paper() -> dict:
    return dict(pmid="", pmcid="", doi="", title="", abstract="", authors=[],
                journal="", year="", pub_types=[], source="", fulltext_url="",
                has_fulltext=False, cited_by=None, pdf_path="")


def search_pubmed(query: str, retrieve: int, log) -> list[dict]:
    """Retrieve up to `retrieve` records. Papers WITHOUT full text are kept."""
    try:
        handle = Entrez.esearch(db="pubmed", term=query,
                                retmax=min(retrieve, 500), sort="relevance")
        rec = Entrez.read(handle); handle.close()
    except Exception as e:
        log(f"PubMed search failed: {e}", "err"); return []
    ids = rec.get("IdList", [])
    log(f"PubMed: {rec.get('Count', '?')} hits, pulling {len(ids)} records")
    if not ids:
        return []

    papers, CHUNK = [], 100
    for i in range(0, len(ids), CHUNK):
        chunk = ids[i:i + CHUNK]
        try:
            h = Entrez.efetch(db="pubmed", id=",".join(chunk),
                              rettype="medline", retmode="text")
            records = list(Medline.parse(h)); h.close()
        except Exception as e:
            log(f"PubMed fetch error: {e}", "err"); continue

        for r in records:
            p = blank_paper()
            p["source"]   = "PubMed"
            p["pmid"]     = r.get("PMID", "")
            p["title"]    = (r.get("TI", "") or "").strip()
            p["abstract"] = (r.get("AB", "") or "").strip()
            p["authors"]  = list(r.get("AU", []) or [])
            p["journal"]  = r.get("TA", "") or r.get("JT", "")
            dp = r.get("DP", "")
            p["year"] = dp.split()[0] if dp else ""
            for aid in r.get("AID", []) or []:
                if "[doi]" in aid:
                    p["doi"] = aid.replace(" [doi]", "").strip()
            pmc = r.get("PMC", "")
            p["pmcid"] = pmc
            p["pub_types"] = list(r.get("PT", []) or [])
            if pmc:
                p["fulltext_url"] = f"https://www.ncbi.nlm.nih.gov/pmc/articles/{pmc}/"
                p["has_fulltext"] = True
            elif p["doi"]:
                p["fulltext_url"] = f"https://doi.org/{p['doi']}"
                p["has_fulltext"] = True
            if p["title"]:
                papers.append(p)
        time.sleep(0.34)   # NCBI courtesy rate
    return papers


def search_europepmc(query: str, retrieve: int, log) -> list[dict]:
    url = "https://www.ebi.ac.uk/europepmc/webservices/rest/search"
    papers, cursor, pulled = [], "*", 0
    while pulled < retrieve:
        page = min(100, retrieve - pulled)
        try:
            r = requests.get(url, timeout=30, headers=UA, params={
                "query": query, "format": "json", "resultType": "core",
                "pageSize": page, "cursorMark": cursor,
            })
            data = r.json()
        except Exception as e:
            log(f"Europe PMC error: {e}", "err"); break
        results = (data.get("resultList") or {}).get("result", []) or []
        if not results:
            break
        for it in results:
            p = blank_paper()
            p["source"]   = "EuropePMC"
            p["pmid"]     = it.get("pmid", "") or ""
            p["pmcid"]    = it.get("pmcid", "") or ""
            p["doi"]      = (it.get("doi", "") or "").lower()
            p["title"]    = (it.get("title", "") or "").strip().rstrip(".")
            p["abstract"] = (it.get("abstractText", "") or "").strip()
            p["journal"]  = ((it.get("journalInfo") or {}).get("journal") or {}).get("title", "") \
                            or it.get("bookOrReportDetails", {}).get("publisher", "")
            p["year"]     = str(it.get("pubYear", "") or "")
            auth          = it.get("authorString", "") or ""
            p["authors"]  = [a.strip() for a in auth.split(",") if a.strip()]
            p["pub_types"] = list((it.get("pubTypeList") or {}).get("pubType", []) or [])
            cites = it.get("citedByCount")
            p["cited_by"] = int(cites) if isinstance(cites, int) else None
            if it.get("isOpenAccess") == "Y" and p["pmcid"]:
                p["fulltext_url"] = f"https://europepmc.org/article/PMC/{p['pmcid']}"
                p["has_fulltext"] = True
            elif p["doi"]:
                p["fulltext_url"] = f"https://doi.org/{p['doi']}"
                p["has_fulltext"] = True
            if p["title"]:
                papers.append(p)
        pulled += len(results)
        cursor = data.get("nextCursorMark") or ""
        if not cursor:
            break
    log(f"Europe PMC: {len(papers)} records")
    return papers


def search_openalex(concepts: list[str], retrieve: int, year_from: str, log) -> list[dict]:
    url = "https://api.openalex.org/works"
    params = {"search": " ".join(concepts), "per-page": min(100, max(1, retrieve)),
              "mailto": "research@local"}
    if year_from.strip().isdigit():
        params["filter"] = f"from_publication_date:{year_from.strip()}-01-01"
    try:
        r = requests.get(url, timeout=30, headers=UA, params=params)
        items = r.json().get("results", []) or []
    except Exception as e:
        log(f"OpenAlex error: {e}", "err"); return []
    papers = []
    for it in items:
        p = blank_paper()
        p["source"]   = "OpenAlex"
        p["title"]    = (it.get("title") or "").strip()
        p["abstract"] = reconstruct_abstract(it.get("abstract_inverted_index"))
        p["doi"]      = (it.get("doi") or "").replace("https://doi.org/", "").lower()
        p["year"]     = str(it.get("publication_year") or "")
        p["cited_by"] = it.get("cited_by_count")
        loc = it.get("primary_location") or {}
        src = loc.get("source") or {}
        p["journal"]  = src.get("display_name", "") or ""
        p["authors"]  = [ (a.get("author") or {}).get("display_name", "")
                          for a in (it.get("authorships") or [])[:8] ]
        ids = it.get("ids") or {}
        if ids.get("pmid"):
            p["pmid"] = str(ids["pmid"]).rsplit("/", 1)[-1]
        oa = it.get("best_oa_location") or {}
        if oa.get("pdf_url"):
            p["fulltext_url"] = oa["pdf_url"]; p["has_fulltext"] = True
        elif p["doi"]:
            p["fulltext_url"] = f"https://doi.org/{p['doi']}"; p["has_fulltext"] = True
        if p["title"]:
            papers.append(p)
    log(f"OpenAlex: {len(papers)} records")
    return papers


def search_arxiv(concepts: list[str], retrieve: int, year_from: str, log) -> list[dict]:
    """Physics, maths, CS, quantitative biology/finance, statistics. Atom feed."""
    import xml.etree.ElementTree as ET
    groups = []
    for c in concepts:
        variants = concept_variants(c)
        groups.append("(" + " OR ".join(f'all:"{v}"' for v in variants if "*" not in v) + ")")
    query = " AND ".join(g for g in groups if g != "()")
    if not query:
        return []
    try:
        r = requests.get("http://export.arxiv.org/api/query", timeout=30, headers=UA,
                         params={"search_query": query, "start": 0,
                                 "max_results": min(200, max(1, retrieve)),
                                 "sortBy": "relevance"})
        root = ET.fromstring(r.text)
    except Exception as e:
        log(f"arXiv error: {e}", "err"); return []

    ns = {"a": "http://www.w3.org/2005/Atom", "arx": "http://arxiv.org/schemas/atom"}
    papers = []
    for entry in root.findall("a:entry", ns):
        def txt(tag, default=""):
            el = entry.find(tag, ns)
            return (el.text or "").strip() if el is not None else default
        p = blank_paper()
        p["source"]   = "arXiv"
        p["title"]    = re.sub(r"\s+", " ", txt("a:title"))
        p["abstract"] = re.sub(r"\s+", " ", txt("a:summary"))
        p["year"]     = txt("a:published")[:4]
        p["authors"]  = [(a.find("a:name", ns).text or "").strip()
                         for a in entry.findall("a:author", ns)
                         if a.find("a:name", ns) is not None][:10]
        doi = entry.find("arx:doi", ns)
        if doi is not None and doi.text:
            p["doi"] = doi.text.strip().lower()
        jref = entry.find("arx:journal_ref", ns)
        prim = entry.find("arx:primary_category", ns)
        p["journal"] = (jref.text.strip() if jref is not None and jref.text
                        else f"arXiv · {prim.get('term')}" if prim is not None else "arXiv")
        for link in entry.findall("a:link", ns):
            if link.get("title") == "pdf":
                p["fulltext_url"] = link.get("href", "")
                p["has_fulltext"] = True
        if not p["fulltext_url"]:
            aid = txt("a:id")
            if aid:
                p["fulltext_url"] = aid
                p["has_fulltext"] = True
        if year_from.strip().isdigit() and p["year"].isdigit():
            if int(p["year"]) < int(year_from.strip()):
                continue
        if p["title"]:
            papers.append(p)
    log(f"arXiv: {len(papers)} records")
    return papers


def search_crossref(concepts: list[str], retrieve: int, year_from: str, log) -> list[dict]:
    """Every discipline with a DOI — the widest net, thinnest abstracts."""
    params = {"query.bibliographic": " ".join(concepts),
              "rows": min(100, max(1, retrieve)),
              "select": ("DOI,title,abstract,container-title,issued,author,"
                         "is-referenced-by-count,type,URL,link"),
              "mailto": "research@local", "sort": "relevance"}
    if year_from.strip().isdigit():
        params["filter"] = f"from-pub-date:{year_from.strip()}-01-01"
    try:
        r = requests.get("https://api.crossref.org/works", timeout=30, headers=UA,
                         params=params)
        items = ((r.json() or {}).get("message") or {}).get("items", []) or []
    except Exception as e:
        log(f"Crossref error: {e}", "err"); return []

    papers = []
    for it in items:
        p = blank_paper()
        p["source"]   = "Crossref"
        titles        = it.get("title") or []
        p["title"]    = re.sub(r"\s+", " ", (titles[0] if titles else "")).strip()
        abstract      = it.get("abstract") or ""
        p["abstract"] = re.sub(r"<[^>]+>", " ", abstract)
        p["abstract"] = re.sub(r"\s+", " ", p["abstract"]).strip()
        cont          = it.get("container-title") or []
        p["journal"]  = cont[0] if cont else (it.get("type", "") or "").replace("-", " ")
        parts         = ((it.get("issued") or {}).get("date-parts") or [[]])[0]
        p["year"]     = str(parts[0]) if parts else ""
        p["doi"]      = (it.get("DOI") or "").lower()
        p["cited_by"] = it.get("is-referenced-by-count")
        p["authors"]  = [" ".join(x for x in (a.get("given"), a.get("family")) if x)
                         for a in (it.get("author") or [])[:10]]
        p["pub_types"] = [it.get("type", "")] if it.get("type") else []
        for link in it.get("link") or []:
            if "pdf" in (link.get("content-type") or ""):
                p["fulltext_url"] = link.get("URL", ""); p["has_fulltext"] = True
        if not p["fulltext_url"] and p["doi"]:
            p["fulltext_url"] = f"https://doi.org/{p['doi']}"
            p["has_fulltext"] = True
        if p["title"]:
            papers.append(p)
    log(f"Crossref: {len(papers)} records")
    return papers


def reconstruct_abstract(inv) -> str:
    if not isinstance(inv, dict):
        return ""
    positions = []
    for word, idxs in inv.items():
        for i in idxs:
            positions.append((i, word))
    positions.sort()
    return " ".join(w for _, w in positions)


def norm_title(t: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (t or "").lower()).strip()


def dedupe(papers: list[dict], log) -> list[dict]:
    """Merge on DOI, then PMID, then normalised title. Richer record wins."""
    out, by_doi, by_pmid, by_title = [], {}, {}, {}

    def better(a, b):
        """Merge b into a, keeping the most complete fields."""
        for k in ("abstract", "doi", "pmid", "pmcid", "journal", "year", "fulltext_url"):
            if not a.get(k) and b.get(k):
                a[k] = b[k]
        if not a.get("authors") and b.get("authors"):
            a["authors"] = b["authors"]
        if a.get("cited_by") is None and b.get("cited_by") is not None:
            a["cited_by"] = b["cited_by"]
        a["has_fulltext"] = a["has_fulltext"] or b["has_fulltext"]
        if b["source"] not in a["source"]:
            a["source"] = a["source"] + "+" + b["source"]
        return a

    for p in papers:
        key_doi   = p["doi"] or None
        key_pmid  = p["pmid"] or None
        key_title = norm_title(p["title"]) or None
        hit = None
        if key_doi and key_doi in by_doi:       hit = by_doi[key_doi]
        elif key_pmid and key_pmid in by_pmid:  hit = by_pmid[key_pmid]
        elif key_title and key_title in by_title: hit = by_title[key_title]
        if hit is not None:
            better(hit, p)
        else:
            out.append(p)
            if key_doi:   by_doi[key_doi] = p
            if key_pmid:  by_pmid[key_pmid] = p
            if key_title: by_title[key_title] = p
    log(f"Deduped {len(papers)} → {len(out)} unique records", "accent")
    return out


# ══════════════════════════════════════════════════════════════════════════════
#  Local pre-rank (BM25) — costs milliseconds, spends no model time
# ══════════════════════════════════════════════════════════════════════════════
def tokens(text: str) -> list[str]:
    return [w for w in re.findall(r"[a-z0-9]+", (text or "").lower())
            if w not in STOP and len(w) > 2]


def rank_papers(papers: list[dict], question: str, concepts: list[str]) -> None:
    """Local relevance ranking. Writes p['prerank'] and p['match'] in place.

    BM25 alone rewards a paper that mentions one concept twenty times. A review
    needs papers that cover EVERY concept, so coverage dominates: the score is
    BM25 scaled by how many of your concepts appear at all, plus bonuses for the
    exact phrase and for hits in the title, where topic words actually mean
    something.
    """
    # per-concept token sets, synonyms included
    groups = []
    for c in concepts:
        variants = concept_variants(c)
        toks = set()
        for v in variants:
            toks |= set(tokens(v.replace("*", "")))
        groups.append({"name": c, "tokens": toks,
                       "phrases": [v.replace("*", "").lower() for v in variants]})

    qterms = tokens(question)
    docs, titles = [], []
    for p in papers:
        titles.append(tokens(p.get("title", "")))
        docs.append(tokens(p.get("title", "") + " " + p.get("abstract", "")))

    N = max(1, len(docs))
    avgdl = max(1.0, sum(len(d) for d in docs) / N)
    df = Counter()
    for d in docs:
        for w in set(d):
            df[w] += 1
    k1, b = 1.5, 0.75

    all_terms = set(qterms)
    for g in groups:
        all_terms |= g["tokens"]

    for p, doc, title in zip(papers, docs, titles):
        tf, tft = Counter(doc), Counter(title)
        dl = max(1, len(doc))
        base = 0.0
        for w in all_terms:
            if w not in tf:
                continue
            idf = math.log(1 + (N - df[w] + 0.5) / (df[w] + 0.5))
            freq = tf[w] + 1.6 * tft[w]          # a title hit counts more
            base += idf * (freq * (k1 + 1)) / (freq + k1 * (1 - b + b * dl / avgdl))

        title_l = (p.get("title", "") or "").lower()
        abs_l   = (p.get("abstract", "") or "").lower()
        hit, phrase_bonus = [], 0.0
        for g in groups:
            where = ""
            if any(ph and ph in title_l for ph in g["phrases"]):
                where, phrase_bonus = "title", phrase_bonus + 3.0
            elif any(ph and ph in abs_l for ph in g["phrases"]):
                where, phrase_bonus = "abstract", phrase_bonus + 1.4
            elif g["tokens"] & set(title):
                where, phrase_bonus = "title terms", phrase_bonus + 1.0
            elif g["tokens"] & set(doc):
                where = "terms"
            if where:
                hit.append((g["name"], where))

        coverage = len(hit) / max(1, len(groups))
        score = base * (0.30 + 0.70 * coverage ** 1.5) + phrase_bonus
        if coverage < 0.5 and len(groups) > 1:
            score *= 0.45                        # off-topic: demoted, never hidden
        if p.get("abstract"):
            score += 0.6
        if (p.get("year") or "").isdigit():
            score += min(1.0, max(0.0, (int(p["year"]) - 2010) / 20))
        if p.get("cited_by"):
            score += min(1.2, math.log10(1 + p["cited_by"]) * 0.4)

        p["prerank"] = round(score, 2)
        p["match"] = {"coverage": round(coverage, 2), "hits": hit,
                      "missing": [g["name"] for g in groups
                                  if g["name"] not in [h[0] for h in hit]]}

    papers.sort(key=lambda x: x.get("prerank", 0), reverse=True)


# ══════════════════════════════════════════════════════════════════════════════
#  Theme clustering — cheap, local, no model
# ══════════════════════════════════════════════════════════════════════════════
def cluster_themes(papers: list[dict], k: int = 6) -> list[str]:
    """Assign p['theme'] from the most distinctive shared terms. Returns names."""
    docs = {id(p): set(tokens(p["title"] + " " + p["abstract"][:400])) for p in papers}
    freq = Counter()
    for s in docs.values():
        freq.update(s)
    candidates = [w for w, n in freq.most_common(60) if 2 <= n <= max(2, len(papers) * 0.6)]
    picked = candidates[:k]
    for p in papers:
        s = docs[id(p)]
        best, best_n = "General", -1
        for w in picked:
            if w in s and freq[w] > best_n:
                best, best_n = w, freq[w]
        p["theme"] = best.title()
    names = sorted({p.get("theme", "General") for p in papers})
    return names


# ══════════════════════════════════════════════════════════════════════════════
#  Ollama — rubric scoring
# ══════════════════════════════════════════════════════════════════════════════
# Rubrics per field. PICO only makes sense in the life sciences; every other
# field gets four criteria that mean something there. Same 0-3 scale, same
# quoted-evidence requirement, so a score always reads the same way.
RUBRICS = {
    "life": [("population",   "POPULATION",     "the population, species or sample studied"),
             ("intervention", "INTERVENTION",   "the intervention, exposure or compared conditions"),
             ("outcome",      "OUTCOME",        "the outcomes actually measured"),
             ("design",       "DESIGN",         "study design and how well it is reported")],
    "computing": [("problem",  "PROBLEM / TASK", "the task or problem addressed"),
                  ("approach", "APPROACH",       "the method, model or system proposed"),
                  ("evaluation", "EVALUATION",   "datasets, baselines and metrics used"),
                  ("rigor",    "REPRODUCIBILITY", "released code or data, ablations, stated limitations")],
    "physical": [("system",  "SYSTEM / MATERIAL", "the material, device or physical system"),
                 ("method",  "METHOD",            "the experimental or computational technique"),
                 ("result",  "MEASUREMENT",       "the quantities measured, derived or predicted"),
                 ("rigor",   "VALIDATION",        "controls, error analysis, reproducibility")],
    "social": [("subject",   "SUBJECT / CONTEXT", "the people, place and period studied"),
               ("framework", "FRAMEWORK",         "the theoretical framing or argument"),
               ("evidence",  "EVIDENCE",          "the data, corpus or sources drawn on"),
               ("method",    "METHOD",            "how the evidence was gathered and analysed")],
    "general": [("topic",    "TOPIC MATCH",  "how directly the subject matches the question"),
                ("approach", "APPROACH",     "what was actually done"),
                ("findings", "FINDINGS",     "what was found or concluded"),
                ("rigor",    "RIGOUR",       "quality of evidence and stated limits")],
}

PROFILE_LABELS = {
    "life":      "Life & health sciences",
    "computing": "Computing & information",
    "physical":  "Physical sciences & engineering",
    "social":    "Social sciences & humanities",
    "general":   "General / interdisciplinary",
}

# Which sources are worth querying per field (PubMed indexes biomedicine only;
# arXiv covers physics/maths/CS; OpenAlex and Crossref cover everything).
PROFILE_SOURCES = {
    "life":      dict(pubmed=True,  epmc=True,  arxiv=False, crossref=False, openalex=True),
    "computing": dict(pubmed=False, epmc=False, arxiv=True,  crossref=True,  openalex=True),
    "physical":  dict(pubmed=False, epmc=False, arxiv=True,  crossref=True,  openalex=True),
    "social":    dict(pubmed=False, epmc=False, arxiv=False, crossref=True,  openalex=True),
    "general":   dict(pubmed=True,  epmc=True,  arxiv=True,  crossref=True,  openalex=True),
}

FIELD_LEXICON = {
    "life": """patient patients clinical disease cancer tumour tumor gene genes protein cell
        cells mouse mice rat cohort trial vaccine drug dose mrna receptor enzyme neuron
        neural cortex microbiome bacteria virus antibody diagnosis therapy surgery symptom
        biomarker mortality morbidity epidemiology prevalence incidence dementia diabetes
        cardiac immune inflammation genome sequencing physiology pharmacokinetics""",
    "computing": """algorithm algorithms dataset datasets benchmark benchmarks accuracy
        transformer neural network training inference gpu latency throughput compiler
        software code repository api llm language model embedding classifier reinforcement
        cryptography blockchain robot robotics computer vision nlp database distributed
        kubernetes runtime bug testing usability interface""",
    "physical": """quantum photon electron spin semiconductor alloy catalyst catalysis polymer
        plasma thermal conductivity spectroscopy diffraction crystal lattice tensile fatigue
        turbulence aerodynamic battery electrode electrolyte solar photovoltaic laser optical
        nanoparticle graphene superconductor reactor combustion seismic climate atmospheric
        geology orbital astronomy galaxy stellar""",
    "social": """policy policies students student teacher curriculum school survey respondents
        participants interview interviews discourse narrative gender race ethnicity migration
        labour labor employment wage inequality poverty urban rural governance democracy
        election parliament colonial archive archival historiography literary rhetoric
        ethnography qualitative econometric welfare household""",
}
FIELD_LEXICON = {k: set(v.split()) for k, v in FIELD_LEXICON.items()}


def detect_field(question: str, terms: list[str] = None) -> str:
    """Guess the field from the question's vocabulary. Ties go to general."""
    words = set(tokens(question)) | set(tokens(" ".join(terms or [])))
    hits = {k: len(words & lex) for k, lex in FIELD_LEXICON.items()}
    best = max(hits, key=hits.get)
    return best if hits[best] >= 1 else "general"


def rubric_for(profile: str):
    return RUBRICS.get(profile) or RUBRICS["general"]


CRITERIA = [(k, label) for k, label, _ in RUBRICS["life"]]   # legacy default

def build_rubric_prompt(paper: dict, question: str, abstract: str, profile: str) -> str:
    spec = rubric_for(profile)
    crit_lines = "\n".join(f"- {k}: {desc}" for k, _label, desc in spec)
    skeleton = ",\n ".join(f'"{k}":{{"score":0,"evidence":""}}' for k, _l, _d in spec)
    return f"""You are screening papers for a systematic literature review in
{PROFILE_LABELS.get(profile, 'this field')}.

Research question: "{question}"

Paper title: {paper.get('title','')}
Venue/year: {paper.get('journal','')} {paper.get('year','')}
Abstract: {abstract}

Score the paper against the research question on four criteria. Each is an integer
0-3: 0 = absent or mismatched, 1 = weak or indirect, 2 = partial match, 3 = direct
match. For each criterion quote a SHORT span (max 15 words) copied verbatim from the
title or abstract as evidence. If the text does not support it, score 0 and leave
evidence empty.

{crit_lines}

Respond with ONLY this JSON object, no markdown, no commentary:
{{{skeleton},
 "paper_type":"a short label for what kind of work this is",
 "summary":"2-3 plain-English sentences: what it investigates, how, main finding"}}"""


def check_ollama(model: str):
    try:
        r = requests.get(f"{OLLAMA_HOST}/api/tags", timeout=5)
        if r.status_code != 200:
            return False, "Ollama returned an error."
        names = [m.get("name", "") for m in r.json().get("models", [])]
        if model in names:
            return True, model
        base = model.split(":")[0]
        near = [n for n in names if n.split(":")[0] == base]
        if near:
            return True, near[0]
        return False, f"Model '{model}' not pulled. Run: ollama pull {model}"
    except requests.exceptions.ConnectionError:
        return False, "Ollama is not running. Start it with: ollama serve"
    except Exception as e:
        return False, str(e)


def _pack(data: dict, profile: str) -> dict:
    """Clamp one model answer into a scored record."""
    crit, total = [], 0
    for key, label, _desc in rubric_for(profile):
        node = data.get(key) or {}
        if isinstance(node, (int, float)):
            node = {"score": node, "evidence": ""}
        try:
            val = int(round(float(node.get("score", 0))))
        except Exception:
            val = 0
        val = max(0, min(3, val))
        ev = str(node.get("evidence", "") or "").strip().strip('"')
        crit.append({"key": key, "label": label, "score": val, "evidence": ev[:180]})
        total += val
    return {"total": total, "criteria": crit, "profile": profile,
            "paper_type": str(data.get("paper_type", "") or "")[:40],
            "summary": str(data.get("summary", "") or "").strip()}


def _json_from(raw: str):
    raw = (raw or "").strip()
    try:
        return json.loads(raw)
    except Exception:
        m = re.search(r"\{[\s\S]*\}", raw)
        if m:
            try:
                return json.loads(m.group())
            except Exception:
                return None
    return None


def score_group(group: list, question: str, model: str, log, fast: bool = False,
                profile: str = "general") -> dict:
    """Score several papers in ONE Ollama call. Returns {list index: result}.

    One call for five papers is far cheaper than five calls: the prompt preamble,
    the model's warm-up and the HTTP round trip are paid once.
    """
    if len(group) == 1:
        r = score_paper(group[0], question, model, log, fast, profile)
        return {0: r} if r else {}

    spec = rubric_for(profile)
    lim  = 650 if fast else 1300
    listing = "\n\n".join(
        f"[{i}] TITLE: {p.get('title','')}\nABSTRACT: {(p.get('abstract') or '')[:lim]}"
        for i, p in enumerate(group, 1))
    crit_lines = "\n".join(f"- {k}: {desc}" for k, _l, desc in spec)
    skel = ",".join(f'"{k}":{{"score":0,"evidence":""}}' for k, _l, _d in spec)
    prompt = (
        f"You are screening papers for a literature review in "
        f"{PROFILE_LABELS.get(profile, 'this field')}.\n\n"
        f'Research question: "{question}"\n\n'
        f"Below are {len(group)} papers. Score EACH against the question on four criteria,\n"
        f"each an integer 0-3 (0 absent/mismatched, 1 weak, 2 partial, 3 direct), quoting a\n"
        f"short verbatim span (max 15 words) from that paper's own title or abstract as\n"
        f"evidence. Never mix papers up; if the text does not support a criterion, score 0\n"
        f"and leave evidence empty.\n\n{crit_lines}\n\n{listing}\n\n"
        f"Respond with ONLY one JSON object keyed by the paper numbers, no commentary:\n"
        f'{{"1":{{{skel},"paper_type":"","summary":"2-3 sentences"}}, … up to "{len(group)}"}}')
    try:
        r = requests.post(f"{OLLAMA_HOST}/api/generate", timeout=600, json={
            "model": model, "prompt": prompt, "stream": False,
            "format": "json", "keep_alive": "10m",
            "options": {"temperature": 0, "top_p": 0.9, "seed": 7,
                        "num_predict": len(group) * (260 if fast else 430)},
        })
        data = _json_from((r.json() or {}).get("response", ""))
    except Exception as e:
        log(f"  batch scoring error: {e}", "err")
        return {}
    if not isinstance(data, dict):
        return {}

    out = {}
    for i in range(1, len(group) + 1):
        node = data.get(str(i)) or data.get(i)
        if isinstance(node, dict) and node:
            out[i - 1] = _pack(node, profile)
    return out


def score_paper(paper: dict, question: str, model: str, log, fast: bool = False,
                profile: str = "general") -> dict:
    """fast=True sends less context and asks for a shorter answer — roughly twice
    the speed, at the cost of blunter evidence quotes."""
    abstract = (paper.get("abstract") or "")[:1200 if fast else 3000]
    prompt = build_rubric_prompt(paper, question, abstract, profile)
    try:
        r = requests.post(f"{OLLAMA_HOST}/api/generate", timeout=240, json={
            "model": model, "prompt": prompt, "stream": False,
            "format": "json", "keep_alive": "10m",
            "options": {"temperature": 0, "top_p": 0.9, "seed": 7,
                        "num_predict": 380 if fast else 700},
        })
        raw = (r.json() or {}).get("response", "").strip()
    except Exception as e:
        log(f"  Ollama error: {e}", "err"); return {}

    data = _json_from(raw)
    if not isinstance(data, dict):
        return {}
    return _pack(data, profile)


def draft_synthesis(papers: list[dict], question: str, model: str, log) -> str:
    """Ask the model for a review paragraph over the INCLUDED papers only."""
    if not papers:
        return ""
    lines = []
    for i, p in enumerate(papers, 1):
        s = p.get("scored") or {}
        # works with or without AI scoring — fall back to the abstract
        gist = s.get("summary") or (p.get("abstract") or "")[:600] or "(no abstract on record)"
        score = f" score {s['total']}/12" if s.get("total") is not None else ""
        lines.append(f"[{i}] {p['title']} ({p.get('journal','')} {p.get('year','')})"
                     f"{score} — {gist}")
    body = "\n".join(lines)[:14000]
    prompt = (
        f'Write the "findings" section of a literature review answering:\n"{question}"\n\n'
        f"Use ONLY these papers and cite them by their bracket numbers.\n\n{body}\n\n"
        "Write 3-5 paragraphs of plain academic prose. Group by theme, name points of "
        "agreement and disagreement, and end with one paragraph on what is still missing. "
        "No bullet lists, no headings, no invented citations."
    )
    try:
        r = requests.post(f"{OLLAMA_HOST}/api/generate", timeout=600, json={
            "model": model, "prompt": prompt, "stream": False, "keep_alive": "10m",
            "options": {"temperature": 0.3, "num_predict": 1400},
        })
        return (r.json() or {}).get("response", "").strip()
    except Exception as e:
        log(f"Synthesis failed: {e}", "err")
        return ""


# ══════════════════════════════════════════════════════════════════════════════
#  PDFs and files
# ══════════════════════════════════════════════════════════════════════════════
def sanitize(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*\n\r]', "_", str(name)).strip()[:60]


def make_folder(topic: str) -> Path:
    folder = BASE_DIR / TODAY / (sanitize(topic) or "review")
    folder.mkdir(parents=True, exist_ok=True)
    return folder


def download_pdf(paper: dict, folder: Path, log) -> str:
    pmc  = paper.get("pmcid", "") or ""
    dest = folder / f"{paper.get('year','')}_{sanitize(paper.get('title','untitled'))}.pdf"
    if dest.exists() and dest.stat().st_size > 10240:
        return str(dest)

    urls = []
    if pmc:
        pid = pmc if pmc.upper().startswith("PMC") else "PMC" + pmc
        urls.append(f"https://www.ncbi.nlm.nih.gov/pmc/articles/{pid}/pdf/")
        urls.append(f"https://europepmc.org/api/fulltextRepo?pprId={pid}&type=FILE&fileName={pid}.pdf")
        urls.append(f"https://europepmc.org/articles/{pid}?pdf=render")
    ft = paper.get("fulltext_url", "")
    if ft.lower().endswith(".pdf"):
        urls.append(ft)

    for url in urls:
        try:
            r = requests.get(url, timeout=45, stream=True, headers=UA, allow_redirects=True)
            ct = r.headers.get("content-type", "")
            if r.status_code == 200 and ("pdf" in ct or "octet" in ct):
                with open(dest, "wb") as f:
                    for chunk in r.iter_content(8192):
                        f.write(chunk)
                if dest.stat().st_size > 10240:
                    return str(dest)
                dest.unlink(missing_ok=True)
        except Exception:
            continue
    return ""


HDR_FONT  = Font(bold=True, color="FFFFFF", size=10, name="Arial")
HDR_FILL  = PatternFill("solid", start_color="1d2d3d")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN      = Side(style="thin", color="cccccc")
CELL_BRD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hdr(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font, cell.fill, cell.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
    ws.row_dimensions[row].height = 28


def _rowstyle(ws, row, ncols, even=False):
    bg = "eef3f8" if even else "ffffff"
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font      = Font(name="Arial", size=9)
        cell.fill      = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border    = CELL_BRD


def _link(cell, url, text=None):
    if url:
        cell.hyperlink = url
        cell.value     = text or url
        cell.font      = Font(color="416180", underline="single", name="Arial", size=9)


def fmt_authors(authors, max_n=5):
    names = list(authors or [])
    if not names:
        return ""
    return ", ".join(names[:max_n]) + (" et al." if len(names) > max_n else "")


def save_workbook(papers: list[dict], folder: Path, topic: str, question: str,
                  query: str, decisions: dict, profile: str = "general",
                  sources: str = "") -> Path:
    wb = Workbook()
    ws = wb.active; ws.title = "Review"
    spec = rubric_for(profile)
    short = [lbl.split(" / ")[0][:9].title() for _k, lbl, _d in spec]
    cols = ["#", "Screen", "Score /12"] + short + [
            "Title", "Authors", "Year", "Venue", "Type", "Cited by", "Source",
            "AI summary"] + [f"Evidence — {lbl.title()}" for _k, lbl, _d in spec] + [
            "Pre-rank", "DOI", "PMID", "Full text", "PDF saved"]
    widths = [4, 10, 9] + [7] * 4 + [50, 30, 6, 26, 20, 9, 14, 60] + [34] * 4 + \
             [9, 32, 11, 40, 34]
    for i, (h, w) in enumerate(zip(cols, widths), 1):
        ws.cell(1, i, h)
        ws.column_dimensions[get_column_letter(i)].width = w
    _hdr(ws, len(cols))
    ws.freeze_panes = "A2"

    fills = {"high": PatternFill("solid", start_color="cfe3f5"),
             "mid":  PatternFill("solid", start_color="e8eef5"),
             "low":  PatternFill("solid", start_color="f1f1f2")}
    doi_col  = len(cols) - 3
    pmid_col = len(cols) - 2

    for idx, p in enumerate(papers, 1):
        s   = p.get("scored") or {}
        cr  = s.get("criteria") or []
        by  = {c.get("key"): c for c in cr if isinstance(c, dict)}
        tot = s.get("total", "")
        scores    = [(by.get(k) or {}).get("score", "") for k, _l, _d in spec]
        evidences = [(by.get(k) or {}).get("evidence", "") for k, _l, _d in spec]
        row = [idx, decisions.get(p["key"], ""), tot] + scores + [
               p.get("title", ""), fmt_authors(p.get("authors")), p.get("year", ""),
               p.get("journal", ""), s.get("paper_type", ""),
               p.get("cited_by") if p.get("cited_by") is not None else "",
               p.get("source", ""), s.get("summary", "")] + evidences + [
               p.get("prerank", ""), p.get("doi", ""), p.get("pmid", ""),
               p.get("fulltext_url", ""), p.get("pdf_path", "")]
        for c, v in enumerate(row, 1):
            ws.cell(idx + 1, c, v)
        ws.row_dimensions[idx + 1].height = 66
        _rowstyle(ws, idx + 1, len(cols), even=(idx % 2 == 0))
        if p.get("doi"):
            _link(ws.cell(idx + 1, doi_col), f"https://doi.org/{p['doi']}", p["doi"])
        if p.get("pmid"):
            _link(ws.cell(idx + 1, pmid_col),
                  f"https://pubmed.ncbi.nlm.nih.gov/{p['pmid']}/", p["pmid"])
        if isinstance(tot, int):
            cell = ws.cell(idx + 1, 3)
            cell.fill = fills["high"] if tot >= 9 else fills["mid"] if tot >= 5 else fills["low"]
            cell.font = Font(bold=True, name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="top")

    ws2 = wb.create_sheet("Search")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 110
    meta = [("Topic", topic), ("Research question", question), ("Date", TODAY),
            ("Field", PROFILE_LABELS.get(profile, profile)),
            ("Rubric", ", ".join(lbl for _k, lbl, _d in spec)),
            ("Query sent", query), ("Records kept", len(papers)),
            ("Scored", sum(1 for p in papers if p.get("scored"))),
            ("Included", sum(1 for v in decisions.values() if v == "include")),
            ("Excluded", sum(1 for v in decisions.values() if v == "exclude")),
            ("PDFs saved", sum(1 for p in papers if p.get("pdf_path"))),
            ("Sources", sources or "—"),
            ("Tool", f"Veronica v{VERSION}")]
    for r, (k, v) in enumerate(meta, 1):
        ws2.cell(r, 1, k).font = Font(bold=True, name="Arial", size=10)
        ws2.cell(r, 2, str(v)).font = Font(name="Arial", size=10)
        ws2.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")

    path = folder / f"Review_{sanitize(topic)}_{TODAY}.xlsx"
    wb.save(path)
    return path




def save_ris(path: Path, papers: list) -> Path:
    """RIS — what Zotero, Mendeley and EndNote all import cleanly."""
    def ty(p):
        t = " ".join(p.get("pub_types") or []).lower()
        if "review" in t:        return "JOUR"
        if "preprint" in t or (p.get("source", "").startswith("arXiv")): return "UNPB"
        if "book" in t:          return "BOOK"
        if "conference" in t or "proceedings" in t: return "CPAPER"
        return "JOUR"

    out = []
    for p in papers:
        out.append(f"TY  - {ty(p)}")
        out.append(f"TI  - {p.get('title','')}")
        for a in (p.get("authors") or []):
            out.append(f"AU  - {a}")
        if p.get("journal"):  out.append(f"JO  - {p['journal']}")
        if p.get("year"):     out.append(f"PY  - {p['year']}")
        if p.get("doi"):      out.append(f"DO  - {p['doi']}")
        if p.get("abstract"): out.append(f"AB  - {p['abstract']}")
        if p.get("pmid"):     out.append(f"AN  - {p['pmid']}")
        if p.get("fulltext_url"): out.append(f"UR  - {p['fulltext_url']}")
        if p.get("pdf_path"): out.append(f"L1  - {p['pdf_path']}")
        s = p.get("scored") or {}
        if s.get("total") is not None:
            out.append(f"N1  - Veronica score {s['total']}/12. {s.get('summary','')}")
        out.append(f"DB  - {p.get('source','')}")
        out.append("ER  - ")
        out.append("")
    path.write_text("\n".join(out), encoding="utf-8")
    return path


def save_bibtex(path: Path, papers: list) -> Path:
    def key(p, i):
        first = (p.get("authors") or ["anon"])[0]
        surname = re.sub(r"[^A-Za-z]", "", first.split()[-1] if first else "anon") or "anon"
        word = next((w for w in tokens(p.get("title", "")) if len(w) > 4), "paper")
        return f"{surname.lower()}{p.get('year','')}{word}{i}"

    def esc(s):
        return str(s or "").replace("{", "(").replace("}", ")").replace("\\", "")

    out = []
    for i, p in enumerate(papers, 1):
        kind = "misc" if (p.get("source", "").startswith("arXiv")) else "article"
        fields = [f"  title = {{{esc(p.get('title'))}}}",
                  f"  author = {{{esc(' and '.join(p.get('authors') or []))}}}",
                  f"  year = {{{esc(p.get('year'))}}}"]
        if p.get("journal"): fields.append(f"  journal = {{{esc(p['journal'])}}}")
        if p.get("doi"):     fields.append(f"  doi = {{{esc(p['doi'])}}}")
        if p.get("fulltext_url"): fields.append(f"  url = {{{esc(p['fulltext_url'])}}}")
        s = p.get("scored") or {}
        if s.get("total") is not None:
            fields.append(f"  note = {{Veronica score {s['total']}/12}}")
        out.append(f"@{kind}{{{key(p, i)},\n" + ",\n".join(fields) + "\n}\n")
    path.write_text("\n".join(out), encoding="utf-8")
    return path


# ==============================================================================
#  Word export - a real .docx, written with the standard library only
# ==============================================================================
def _x(s) -> str:
    return (str(s if s is not None else "")
            .replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;"))


class Docx:
    """Minimal WordprocessingML writer: headings, paragraphs, tables,
    external hyperlinks and an updatable table of contents."""

    ACCENT = "416180"
    RULE   = "C9C9CC"
    ZEBRA  = "F1F4F7"
    HEAD   = "E4EAF1"

    def __init__(self, title="Literature review"):
        self.body = []
        self.rels = []
        self.title = title

    def _link(self, url):
        rid = f"rId{100 + len(self.rels)}"
        self.rels.append((rid, url))
        return rid

    def _run(self, text, bold=False, italic=False, size=20, color=None,
             font="Georgia", caps=False):
        rpr = [f'<w:rFonts w:ascii="{font}" w:hAnsi="{font}"/>']
        if bold:   rpr.append("<w:b/>")
        if italic: rpr.append("<w:i/>")
        if caps:   rpr.append("<w:caps/>")
        if color:  rpr.append(f'<w:color w:val="{color}"/>')
        if caps:   rpr.append('<w:spacing w:val="20"/>')
        rpr.append(f'<w:sz w:val="{size}"/><w:szCs w:val="{size}"/>')
        out = ""
        for i, chunk in enumerate(str(text).split("\n")):
            if i:
                out += "<w:br/>"
            out += f'<w:t xml:space="preserve">{_x(chunk)}</w:t>'
        return f"<w:r><w:rPr>{''.join(rpr)}</w:rPr>{out}</w:r>"

    def _hyperlink(self, text, url, size=20):
        rid = self._link(url)
        return (f'<w:hyperlink r:id="{rid}">'
                f'<w:r><w:rPr><w:rStyle w:val="Hyperlink"/>'
                f'<w:rFonts w:ascii="Georgia" w:hAnsi="Georgia"/>'
                f'<w:sz w:val="{size}"/></w:rPr>'
                f'<w:t xml:space="preserve">{_x(text)}</w:t></w:r></w:hyperlink>')

    def para(self, text="", style=None, bold=False, italic=False, size=20,
             color=None, space_after=120, align=None, font="Georgia", caps=False,
             rule_below=False, indent=0):
        ppr = []
        if style:
            ppr.append(f'<w:pStyle w:val="{style}"/>')
        if rule_below:
            ppr.append(f'<w:pBdr><w:bottom w:val="single" w:sz="6" '
                       f'w:color="{self.RULE}"/></w:pBdr>')
        ppr.append(f'<w:spacing w:after="{space_after}" w:line="276" w:lineRule="auto"/>')
        if indent:
            ppr.append(f'<w:ind w:left="{indent}"/>')
        if align:
            ppr.append(f'<w:jc w:val="{align}"/>')
        run = self._run(text, bold, italic, size, color, font, caps) if text != "" else ""
        self.body.append(f"<w:p><w:pPr>{''.join(ppr)}</w:pPr>{run}</w:p>")

    def heading(self, text, level=1):
        self.para(text, style=f"Heading{level}", bold=True,
                  size=34 if level == 1 else 26, color=self.ACCENT,
                  space_after=140, rule_below=(level == 1))

    def title_block(self, title, subtitle=None):
        self.para(title, style="Title", bold=True, size=52, color="1D1F20", space_after=60)
        if subtitle:
            self.para(subtitle, italic=True, size=24, color="424244", space_after=260)

    def toc(self):
        self.body.append(
            '<w:p><w:pPr><w:spacing w:after="200"/></w:pPr>'
            '<w:r><w:fldChar w:fldCharType="begin" w:dirty="true"/></w:r>'
            '<w:r><w:instrText xml:space="preserve"> TOC \\o "1-2" \\h \\z \\u </w:instrText></w:r>'
            '<w:r><w:fldChar w:fldCharType="separate"/></w:r>'
            '<w:r><w:rPr><w:rFonts w:ascii="Georgia" w:hAnsi="Georgia"/>'
            '<w:sz w:val="18"/><w:color w:val="7A7A7D"/></w:rPr>'
            '<w:t>Right-click and choose "Update field" to build the contents.</w:t></w:r>'
            '<w:r><w:fldChar w:fldCharType="end"/></w:r></w:p>')

    def page_break(self):
        self.body.append('<w:p><w:r><w:br w:type="page"/></w:r></w:p>')

    def table(self, headers, rows, widths, caption=None, header_size=16, body_size=17):
        """rows: list of lists; a cell is a string, or (text, url) for a link."""
        total = sum(widths) or 1
        grid = "".join(f'<w:gridCol w:w="{int(9360 * w / total)}"/>' for w in widths)
        borders = (f'<w:tblBorders>'
                   f'<w:top w:val="single" w:sz="4" w:color="{self.RULE}"/>'
                   f'<w:bottom w:val="single" w:sz="4" w:color="{self.RULE}"/>'
                   f'<w:insideH w:val="single" w:sz="2" w:color="{self.RULE}"/>'
                   f'</w:tblBorders>')
        out = [f'<w:tbl><w:tblPr><w:tblW w:w="5000" w:type="pct"/>{borders}'
               f'<w:tblLayout w:type="fixed"/>'
               f'<w:tblLook w:val="04A0" w:firstRow="1" w:lastRow="0" w:firstColumn="1" '
               f'w:lastColumn="0" w:noHBand="0" w:noVBand="1"/></w:tblPr>'
               f'<w:tblGrid>{grid}</w:tblGrid>']

        def cell(content, w, shade=None, bold=False, size=body_size, caps=False):
            wpx = int(9360 * w / total)
            shading = f'<w:shd w:val="clear" w:color="auto" w:fill="{shade}"/>' if shade else ""
            if isinstance(content, tuple):
                inner = self._hyperlink(content[0], content[1], size)
            else:
                inner = self._run(content, bold=bold, size=size,
                                  color=self.ACCENT if caps else None, caps=caps)
            return (f'<w:tc><w:tcPr><w:tcW w:w="{wpx}" w:type="dxa"/>{shading}'
                    f'<w:tcMar><w:top w:w="70" w:type="dxa"/><w:bottom w:w="70" w:type="dxa"/>'
                    f'<w:left w:w="90" w:type="dxa"/><w:right w:w="90" w:type="dxa"/></w:tcMar>'
                    f'<w:vAlign w:val="top"/></w:tcPr>'
                    f'<w:p><w:pPr><w:spacing w:after="0" w:line="240" w:lineRule="auto"/>'
                    f'</w:pPr>{inner}</w:p></w:tc>')

        if headers:
            cells = "".join(cell(h, w, self.HEAD, bold=True, size=header_size, caps=True)
                            for h, w in zip(headers, widths))
            out.append(f'<w:tr><w:trPr><w:tblHeader/></w:trPr>{cells}</w:tr>')
        for i, row in enumerate(rows):
            shade = self.ZEBRA if i % 2 else None
            cells = "".join(cell(c, w, shade) for c, w in zip(row, widths))
            out.append(f"<w:tr>{cells}</w:tr>")
        out.append("</w:tbl>")
        self.body.append("".join(out))
        if caption:
            self.para(caption, size=16, color="7A7A7D", space_after=240)
        else:
            self.para("", space_after=200)

    def save(self, path: Path):
        doc = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
               '<w:document '
               'xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main" '
               'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
               "<w:body>" + "".join(self.body) +
               '<w:sectPr><w:pgSz w:w="11906" w:h="16838"/>'
               '<w:pgMar w:top="1134" w:right="1021" w:bottom="1134" w:left="1021" '
               'w:header="709" w:footer="709" w:gutter="0"/></w:sectPr>'
               "</w:body></w:document>")

        rels = "".join(
            f'<Relationship Id="{rid}" '
            f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink" '
            f'Target="{_x(url)}" TargetMode="External"/>' for rid, url in self.rels)

        styles = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                  '<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
                  '<w:docDefaults><w:rPrDefault><w:rPr>'
                  '<w:rFonts w:ascii="Georgia" w:hAnsi="Georgia"/>'
                  '<w:sz w:val="20"/><w:szCs w:val="20"/></w:rPr></w:rPrDefault>'
                  '<w:pPrDefault><w:pPr><w:spacing w:after="120" w:line="276" '
                  'w:lineRule="auto"/></w:pPr></w:pPrDefault></w:docDefaults>'
                  '<w:style w:type="paragraph" w:default="1" w:styleId="Normal">'
                  '<w:name w:val="Normal"/></w:style>'
                  '<w:style w:type="paragraph" w:styleId="Title">'
                  '<w:name w:val="Title"/><w:basedOn w:val="Normal"/>'
                  '<w:pPr><w:outlineLvl w:val="0"/></w:pPr></w:style>'
                  '<w:style w:type="paragraph" w:styleId="Heading1">'
                  '<w:name w:val="heading 1"/><w:basedOn w:val="Normal"/>'
                  '<w:pPr><w:keepNext/><w:outlineLvl w:val="0"/>'
                  '<w:spacing w:before="320" w:after="140"/></w:pPr></w:style>'
                  '<w:style w:type="paragraph" w:styleId="Heading2">'
                  '<w:name w:val="heading 2"/><w:basedOn w:val="Normal"/>'
                  '<w:pPr><w:keepNext/><w:outlineLvl w:val="1"/>'
                  '<w:spacing w:before="240" w:after="100"/></w:pPr></w:style>'
                  '<w:style w:type="character" w:styleId="Hyperlink">'
                  '<w:name w:val="Hyperlink"/><w:rPr>'
                  f'<w:color w:val="{self.ACCENT}"/><w:u w:val="single"/></w:rPr></w:style>'
                  "</w:styles>")

        ct = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
              '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
              '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
              '<Default Extension="xml" ContentType="application/xml"/>'
              '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
              '<Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>'
              '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
              "</Types>")

        pkg_rels = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                    '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
                    '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
                    "</Relationships>")

        doc_rels = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                    '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
                    + rels + "</Relationships>")

        now = datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%SZ")
        core = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<cp:coreProperties '
                'xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
                'xmlns:dc="http://purl.org/dc/elements/1.1/" '
                'xmlns:dcterms="http://purl.org/dc/terms/" '
                'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
                f"<dc:title>{_x(self.title)}</dc:title>"
                f"<dc:creator>Veronica v{VERSION}</dc:creator>"
                f'<dcterms:created xsi:type="dcterms:W3CDTF">{now}</dcterms:created>'
                "</cp:coreProperties>")

        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
            z.writestr("[Content_Types].xml", ct)
            z.writestr("_rels/.rels", pkg_rels)
            z.writestr("docProps/core.xml", core)
            z.writestr("word/document.xml", doc)
            z.writestr("word/styles.xml", styles)
            z.writestr("word/_rels/document.xml.rels", doc_rels)
        return path


def save_review_docx(path: Path, question: str, prose: str, included: list,
                     excluded: list, profile: str, meta: dict) -> Path:
    spec = rubric_for(profile)
    d = Docx("Literature review - findings")

    d.title_block("Literature review", question or "-")
    d.table(["Field", "Value"],
            [["Date", TODAY],
             ["Discipline", PROFILE_LABELS.get(profile, profile)],
             ["Rubric", ", ".join(lbl for _k, lbl, _d in spec)],
             ["Sources searched", meta.get("sources", "-")],
             ["Records retrieved", str(meta.get("retrieved", "-"))],
             ["AI-scored", str(meta.get("scored", 0))],
             ["Included", str(len(included))],
             ["Screened out", str(len(excluded))],
             ["Search query", meta.get("query", "-")],
             ["Tool", f"Veronica v{VERSION} - model {meta.get('model', '-')}"]],
            widths=[26, 74])

    d.heading("Contents")
    d.toc()
    d.page_break()

    d.heading("Findings")
    for block in [b.strip() for b in (prose or "").split("\n") if b.strip()]:
        d.para(block, size=21, space_after=180)
    d.para("Drafted from the included papers below; bracket numbers refer to that table. "
           "Verify every claim against the source before use.",
           italic=True, size=17, color="7A7A7D", space_after=240)

    d.heading("Included papers")
    rows = []
    for i, p in enumerate(included, 1):
        s = p.get("scored") or {}
        url = (f"https://doi.org/{p['doi']}" if p.get("doi") else p.get("fulltext_url") or "")
        title = p.get("title", "")
        rows.append([str(i), (title, url) if url else title,
                     f"{p.get('journal','')} {p.get('year','')}".strip(),
                     "" if p.get("cited_by") is None else str(p["cited_by"]),
                     f"{s['total']}/12" if s.get("total") is not None else "-"])
    d.table(["#", "Title", "Venue / year", "Cited", "Score"], rows, [4, 44, 26, 8, 8],
            caption="Titles link to the DOI or full text.")

    d.heading("Summaries and evidence")
    for i, p in enumerate(included, 1):
        s = p.get("scored") or {}
        d.para(f"[{i}] {p.get('title','')}", bold=True, size=20, space_after=40)
        d.para(f"{fmt_authors(p.get('authors'), 4)} - {p.get('journal','')} "
               f"{p.get('year','')}", size=17, color="7A7A7D", space_after=60)
        gist = s.get("summary") or (p.get("abstract") or "")[:900] or "No abstract on record."
        d.para(gist, size=19, space_after=100)
        crit = s.get("criteria") or []
        if isinstance(crit, dict):
            crit = [{"label": k.upper(), "score": v.get("score", 0),
                     "evidence": v.get("evidence", "")} for k, v in crit.items()]
        if crit:
            d.table(["Criterion", "Score", "Evidence quoted from the abstract"],
                    [[c.get("label", ""), f"{c.get('score', 0)}/3",
                      '"' + str(c.get("evidence")) + '"' if c.get("evidence") else "-"]
                     for c in crit], [24, 8, 68])
        else:
            d.para("Not AI-scored - ranked locally against the question.",
                   italic=True, size=16, color="7A7A7D", space_after=200)

    if excluded:
        d.heading("Screened out")
        d.table(["Title", "Venue / year", "Score"],
                [[p.get("title", ""),
                  f"{p.get('journal','')} {p.get('year','')}".strip(),
                  f"{(p.get('scored') or {}).get('total')}/12"
                  if (p.get("scored") or {}).get("total") is not None else "-"]
                 for p in excluded], [62, 28, 10])

    d.heading("References")
    for i, p in enumerate(included, 1):
        tail = f"{p.get('journal','')} {p.get('year','')}".strip()
        d.para(f"[{i}] {fmt_authors(p.get('authors'), 6)}. {p.get('title','')}. {tail}.",
               size=18, space_after=40)
        if p.get("doi"):
            link = "https://doi.org/" + p["doi"]
            d.body.append('<w:p><w:pPr><w:spacing w:after="120"/><w:ind w:left="340"/>'
                          '</w:pPr>' + d._hyperlink(link, link, 17) + "</w:p>")

    d.heading("Method")
    scored_note = (f"The top {meta.get('scored', 0)} were scored by "
                   f"{meta.get('model', 'a local model')} on the "
                   f"{PROFILE_LABELS.get(profile, profile).lower()} rubric "
                   f"({', '.join(lbl.lower() for _k, lbl, _d in spec)}), each criterion 0-3 "
                   f"with a verbatim evidence span, totalling 0-12. "
                   if meta.get("scored") else
                   "No AI scoring was applied; ordering is local relevance only. ")
    d.para(f"Records were retrieved from {meta.get('sources', '-')} using the query recorded "
           f"above, deduplicated on DOI, PMID and normalised title, then ranked locally with "
           f"BM25 against the research question. " + scored_note +
           "Screening decisions were made by the author. Everything ran on the author's "
           "machine; nothing left it except the literature searches themselves.",
           size=19, space_after=200)

    return d.save(path)


# ══════════════════════════════════════════════════════════════════════════════
#  Deriving search terms from the question
# ══════════════════════════════════════════════════════════════════════════════
QUESTION_WORDS = set("""which what how why when where who whom whose does do did is are was
were can could should would will has have had any many much most best better effect effects
impact role use used using study studies studied research investigate investigated compare
compared comparison evidence known about there their between among within across for from
with without into onto over under than then also more less least very such other others
paper papers article articles literature review reviews""".split())

PLURAL_FIX = {"models": "model", "mice": "mice", "studies": "study", "trials": "trial",
              "patients": "patients", "children": "children", "outcomes": "outcome",
              "therapies": "therapy", "diseases": "disease", "cells": "cell",
              "drugs": "drug", "methods": "method", "assays": "assay", "rats": "rats"}


def _singular(word: str) -> str:
    w = word.lower()
    if w in PLURAL_FIX:
        return PLURAL_FIX[w]
    if len(w) > 4 and w.endswith("ies"):
        return w[:-3] + "y"
    if len(w) > 3 and w.endswith("s") and not w.endswith("ss") and not w.endswith("us"):
        return w[:-1]
    return w


def derive_terms(question: str, max_terms: int = 4) -> list[str]:
    """Turn a plain-English question into the concepts we search on.

    Contiguous runs of meaningful words become one concept; known phrases in
    SYNONYMS win outright. This is what the search is actually built from, so
    the UI shows the result and lets the researcher edit it.
    """
    q = (question or "").lower()
    q = q.replace("’", "'")
    found, taken = [], set()

    # 1 — known phrases first, longest wins
    for phrase in sorted(SYNONYMS.keys(), key=len, reverse=True):
        needle = phrase.replace("'", "'")
        if needle in q and not any(needle in t for t in found):
            found.append(phrase)
            taken.update(needle.split())

    # 2 — contiguous runs of content words
    words = re.findall(r"[a-z0-9'\-]+", q)
    run, runs = [], []
    for w in words:
        if w in STOP or w in QUESTION_WORDS or len(w) < 3:
            if run:
                runs.append(run); run = []
        else:
            run.append(w)
    if run:
        runs.append(run)

    for r in runs:
        if any(w in taken for w in r):
            continue
        phrase = " ".join(r[:-1] + [_singular(r[-1])]) if len(r) > 1 else _singular(r[0])
        phrase = phrase.strip("'- ")
        if phrase and phrase not in found:
            found.append(phrase)

    # longest/most specific first, then cap
    found.sort(key=lambda t: (-len(t.split()), -len(t)))
    return found[:max_terms]


# ══════════════════════════════════════════════════════════════════════════════
#  GUI helpers
# ══════════════════════════════════════════════════════════════════════════════
class ScrollFrame(tk.Frame):
    """A vertically scrollable container that behaves on every platform."""

    def __init__(self, parent, bg):
        super().__init__(parent, bg=bg)
        self.canvas = tk.Canvas(self, bg=bg, highlightthickness=0, bd=0)
        self.vbar   = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.inner  = tk.Frame(self.canvas, bg=bg)
        self.canvas.configure(yscrollcommand=self.vbar.set)
        self.vbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        self._win = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.inner.bind("<Configure>", self._on_inner)
        self.canvas.bind("<Configure>", self._on_canvas)
        for w in (self.canvas, self.inner):
            w.bind("<Enter>", lambda e: self._bind_wheel(True))
            w.bind("<Leave>", lambda e: self._bind_wheel(False))

    def _on_inner(self, _e=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas(self, e):
        self.canvas.itemconfigure(self._win, width=e.width)

    def _bind_wheel(self, on):
        if on:
            self.canvas.bind_all("<MouseWheel>", self._wheel)
            self.canvas.bind_all("<Button-4>", lambda e: self.canvas.yview_scroll(-2, "units"))
            self.canvas.bind_all("<Button-5>", lambda e: self.canvas.yview_scroll(2, "units"))
        else:
            self.canvas.unbind_all("<MouseWheel>")
            self.canvas.unbind_all("<Button-4>")
            self.canvas.unbind_all("<Button-5>")

    def _wheel(self, e):
        step = -1 * (e.delta // 120 if abs(e.delta) >= 120 else e.delta)
        self.canvas.yview_scroll(step, "units")

    def retheme(self, bg):
        self.configure(bg=bg)
        self.canvas.configure(bg=bg)
        self.inner.configure(bg=bg)


# ══════════════════════════════════════════════════════════════════════════════
#  GUI
# ══════════════════════════════════════════════════════════════════════════════
class Veronica(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"Veronica · AI Literature Review · v{VERSION}")

        self._load_settings()
        self.T          = THEMES[self.mode]
        self._reg       = []
        self.papers     = []
        self.view       = []
        self.decisions  = {}
        self.themes     = []
        self.terms      = list(self.settings.get("terms") or [])
        self.folder     = None
        self.running    = False
        self.fullscreen = False
        self.msgq       = queue.Queue()
        self.selected   = None
        self.scoring    = False
        self.ctx        = None
        self.profile    = "general"
        self.theme_filter = tk.StringVar(value="All themes")

        self._size_to_screen()
        self._set_icon()
        self._make_vars()
        self._build()
        self._apply_theme()
        self._render_chips()
        self.after(120, self._drain)
        self.after(400, self.check_model)
        if self.v_field.get() != "auto":
            self.apply_profile(self.v_field.get())
        self.protocol("WM_DELETE_WINDOW", self._quit)
        self.log(f"Veronica v{VERSION} — PubMed + Europe PMC + OpenAlex + Ollama", "accent")
        self.log(f"Output folder: {BASE_DIR}")
        self.log("Type a research question, check the search terms, press Run review.", "dim")
        if not self.settings.get("seen_intro"):
            self.after(700, self._first_run)

    # ── first run ────────────────────────────────────────────────────────────
    def _first_run(self):
        T = self.T
        win = tk.Toplevel(self)
        win.title("Welcome to Veronica")
        win.transient(self)
        w, h = int(560 * self.scale), int(470 * self.scale)
        win.geometry(f"{w}x{h}+{self.winfo_rootx() + 80}+{self.winfo_rooty() + 70}")
        win.configure(bg=T["bg"])

        tk.Label(win, text="VERONICA", anchor="w", bg=T["bg"], fg=T["accent"],
                 font=self.fs(22, "bold", "head")).pack(fill="x", padx=26, pady=(24, 2))
        tk.Label(win, text="Literature review, on your machine.", anchor="w",
                 bg=T["bg"], fg=T["mut"],
                 font=self.fs(12)).pack(fill="x", padx=26, pady=(0, 16))

        for n, title, text in (
            ("1", "Ask a real question",
             "Veronica reads the concepts out of it and searches up to five databases. "
             "Ranking is local and takes seconds."),
            ("2", "Screen what came back",
             "I / X on a row includes or excludes it. Your decisions drive the Excel "
             "workbook and the Word draft."),
            ("3", "Spend the model only if you want to",
             "AI scoring adds a 0–12 rubric with quoted evidence, at a few seconds a "
             "paper. The switch is in the main window; OFF is a valid way to work.")):
            row = tk.Frame(win, bg=T["bg"]); row.pack(fill="x", padx=26, pady=6)
            tk.Label(row, text=n, bg=T["accent"], fg=T["on_accent"], width=3,
                     font=self.fs(13, "bold", "head")).pack(side="left", padx=(0, 12))
            col = tk.Frame(row, bg=T["bg"]); col.pack(side="left", fill="x", expand=True)
            tk.Label(col, text=title, anchor="w", bg=T["bg"], fg=T["text"],
                     font=self.fs(12, "bold")).pack(fill="x")
            tk.Label(col, text=text, anchor="w", justify="left", bg=T["bg"], fg=T["mut"],
                     wraplength=int(420 * self.scale),
                     font=self.fs(10)).pack(fill="x")

        status = tk.Label(win, text="checking Ollama…", anchor="w", bg=T["bg"],
                          fg=T["mut2"], justify="left", wraplength=int(500 * self.scale),
                          font=self.fs(10, kind="mono"))
        status.pack(fill="x", padx=26, pady=(18, 0))

        def probe():
            ok, msg = check_ollama(self.v_smodel.get().strip() or SCORE_MODEL)
            if ok:
                self.after(0, lambda: status.configure(
                    text=f"● Ollama ready · {msg} — scoring and drafting available.",
                    fg=T["ok"]))
            else:
                self.after(0, lambda: status.configure(
                    text=f"○ {msg}\nSearching, ranking, screening and Excel all work "
                         f"without it. Install from ollama.com when you want scoring "
                         f"or drafting.", fg=T["warn"]))
        threading.Thread(target=probe, daemon=True).start()

        btns = tk.Frame(win, bg=T["bg"]); btns.pack(fill="x", padx=26, pady=22)

        def close(sample=False):
            self.settings["seen_intro"] = True
            if sample and not self.txt_q.get("1.0", "end").strip():
                self.txt_q.insert("1.0", "Which mouse models are used to study mild "
                                         "cognitive impairment in Alzheimer's disease?")
                self.refresh_terms(force=True)
            self._save_settings()
            win.destroy()

        tk.Button(btns, text="TRY A SAMPLE QUESTION", relief="flat", bd=0, cursor="hand2",
                  padx=12, pady=6, bg=T["bg"], fg=T["text"], highlightthickness=1,
                  highlightbackground=T["line"], activeforeground=T["accent"],
                  font=self.fs(10, kind="mono"),
                  command=lambda: close(True)).pack(side="left")
        tk.Button(btns, text="START", relief="flat", bd=0, cursor="hand2", padx=22, pady=6,
                  bg=T["accent"], fg=T["on_accent"], activebackground=T["accent_dim"],
                  font=self.fs(12, "bold", "head"),
                  command=lambda: close(False)).pack(side="right")

    # ── window sizing ─────────────────────────────────────────────────────────
    def _size_to_screen(self):
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.sw, self.sh = sw, sh
        # font scale from the screen's short side, so text stays legible on a
        # 1366x768 laptop and doesn't look tiny on a 4K panel
        self.scale = 0.9 if sh < 800 else 1.0 if sh < 1100 else 1.15 if sh < 1500 else 1.3
        w = min(int(sw * 0.94), 1760)
        h = min(int(sh * 0.90), 1160)
        x, y = max(0, (sw - w) // 2), max(0, (sh - h) // 2 - 12)
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.minsize(min(980, sw - 40), min(620, sh - 80))
        self.bind("<F11>", lambda e: self.toggle_fullscreen())
        self.bind("<Escape>", lambda e: self.exit_fullscreen())
        self.bind("<Command-Control-f>", lambda e: self.toggle_fullscreen())
        if self.settings.get("maximized"):
            self.after(60, self.maximize)

    def _set_icon(self):
        here = Path(__file__).resolve().parent
        ico = here / "assets" / "veronica.ico"
        png = here / "assets" / "veronica_icon_256.png"
        if sys.platform == "win32" and ico.exists():
            try:
                self.iconbitmap(default=str(ico))
                return
            except tk.TclError:
                pass
        if png.exists():
            try:
                self._icon = tk.PhotoImage(file=str(png))
                self.iconphoto(True, self._icon)
            except tk.TclError:
                pass

    def fs(self, size, weight="normal", kind="body"):
        fam = {"body": self.T["body"], "mono": self.T["mono"], "head": self.T["head"]}[kind]
        fb  = {"body": ["Barlow", "Georgia", "Helvetica", "TkDefaultFont"],
               "mono": ["Menlo", "Consolas", "DejaVu Sans Mono", "Courier"],
               "head": ["Barlow Condensed", "Georgia", "Helvetica"]}[kind]
        return font_or(fam, fb, max(8, int(round(size * self.scale))), weight)

    def maximize(self):
        try:
            self.state("zoomed")                      # Windows / most Linux WMs
        except tk.TclError:
            try:
                self.attributes("-zoomed", True)      # some X11 WMs
            except tk.TclError:
                self.geometry(f"{self.sw}x{self.sh - 40}+0+0")

    def toggle_fullscreen(self):
        self.fullscreen = not self.fullscreen
        try:
            self.attributes("-fullscreen", self.fullscreen)
        except tk.TclError:
            self.maximize()
        self.btn_full.configure(text="EXIT FULL SCREEN" if self.fullscreen else "FULL SCREEN")

    def exit_fullscreen(self):
        if self.fullscreen:
            self.toggle_fullscreen()

    def _quit(self):
        self._save_settings()
        self.destroy()

    # ── settings ──────────────────────────────────────────────────────────────
    def _load_settings(self):
        self.settings = {}
        self.mode = "light"
        try:
            self.settings = json.loads(SETTINGS_PATH.read_text())
            self.mode = self.settings.get("mode", "light")
        except Exception:
            pass

    def _make_vars(self):
        s = self.settings
        self.v_retrieve = tk.StringVar(value=str(s.get("retrieve", 200)))
        self.v_keep     = tk.StringVar(value=str(s.get("keep", 40)))
        self.v_year     = tk.StringVar(value=str(s.get("year", "")))
        self.v_workers  = tk.StringVar(value=str(s.get("workers", 4)))
        self.v_model    = tk.StringVar(value=s.get("model", DEFAULT_MODEL))
        self.v_smodel   = tk.StringVar(value=s.get("score_model", SCORE_MODEL))
        self.v_batch    = tk.StringVar(value=str(s.get("batch", 5)))
        self.v_epmc     = tk.BooleanVar(value=s.get("epmc", True))
        self.v_pubmed   = tk.BooleanVar(value=s.get("pubmed", True))
        self.v_oa       = tk.BooleanVar(value=s.get("openalex", True))
        self.v_pdf      = tk.BooleanVar(value=s.get("pdf", False))
        self.v_nofl     = tk.BooleanVar(value=s.get("no_fulltext", True))
        mode = {"Ask": "ask", "Always": "auto", "Never": "off"}.get(
            s.get("score_mode", "ask"), s.get("score_mode", "ask"))
        self.v_score    = tk.StringVar(value=mode if mode in ("ask", "auto", "off") else "ask")
        self.v_fast     = tk.BooleanVar(value=s.get("fast", True))
        self.v_arxiv    = tk.BooleanVar(value=s.get("arxiv", False))
        self.v_cross    = tk.BooleanVar(value=s.get("crossref", False))
        self.v_field    = tk.StringVar(value=s.get("field", "auto"))

    def _save_settings(self):
        try:
            SETTINGS_PATH.write_text(json.dumps({
                "mode": self.mode,
                "question": self.txt_q.get("1.0", "end").strip(),
                "terms": self.terms,
                "retrieve": self.v_retrieve.get(), "keep": self.v_keep.get(),
                "year": self.v_year.get(), "workers": self.v_workers.get(),
                "model": self.v_model.get(), "epmc": self.v_epmc.get(),
                "score_model": self.v_smodel.get(), "batch": self.v_batch.get(),
                "openalex": self.v_oa.get(), "pdf": self.v_pdf.get(),
                "pubmed": self.v_pubmed.get(),
                "no_fulltext": self.v_nofl.get(),
                "score_mode": self.v_score.get(), "fast": self.v_fast.get(),
                "arxiv": self.v_arxiv.get(), "crossref": self.v_cross.get(),
                "field": self.v_field.get(),
                "seen_intro": bool(self.settings.get("seen_intro")),
                "maximized": bool(self.settings.get("maximized")),
            }, indent=1))
        except Exception:
            pass

    # ── theming ───────────────────────────────────────────────────────────────
    def reg(self, w, role):
        self._reg.append((w, role)); return w

    def _apply_theme(self):
        T = self.T = THEMES[self.mode]
        self.configure(bg=T["bg"])
        body, mono, small = self.fs(11), self.fs(10, kind="mono"), self.fs(9, kind="mono")
        for w, role in list(self._reg):
            try:
                if   role == "bg":       w.configure(bg=T["bg"])
                elif role == "surface":  w.configure(bg=T["surface"])
                elif role == "line":     w.configure(bg=T["line"])
                elif role == "text":     w.configure(bg=T["bg"], fg=T["text"], font=body)
                elif role == "text_s":   w.configure(bg=T["surface"], fg=T["text"], font=body)
                elif role == "brand":    w.configure(bg=T["surface"], fg=T["accent"],
                                                     font=self.fs(17, "bold", "head"))
                elif role == "label":    w.configure(bg=T["bg"], fg=T["mut2"], font=small)
                elif role == "label_s":  w.configure(bg=T["surface"], fg=T["mut2"], font=small)
                elif role == "mono":     w.configure(bg=T["bg"], fg=T["mut"], font=small)
                elif role == "mono_s":   w.configure(bg=T["surface"], fg=T["mut"], font=small)
                elif role == "hint":     w.configure(bg=T["bg"], fg=T["mut2"], font=self.fs(10))
                elif role == "entry":
                    w.configure(bg=T["surface"], fg=T["text"], insertbackground=T["accent"],
                                highlightbackground=T["line"], highlightcolor=T["accent"],
                                font=self.fs(13))
                elif role == "query":
                    w.configure(bg=T["surface"], fg=T["mut"], insertbackground=T["accent"],
                                highlightbackground=T["line"], font=small)
                elif role == "primary":
                    w.configure(bg=T["accent"], fg=T["on_accent"], activebackground=T["accent_dim"],
                                activeforeground=T["on_accent"], font=self.fs(13, "bold", "head"))
                elif role == "ghost":
                    w.configure(bg=T["bg"], fg=T["text"], activebackground=T["surface"],
                                activeforeground=T["accent"], highlightbackground=T["line"],
                                font=small)
                elif role == "ghost_s":
                    w.configure(bg=T["surface"], fg=T["text"], activebackground=T["surface2"],
                                activeforeground=T["accent"], highlightbackground=T["line"],
                                font=small)
                elif role == "chip":
                    w.configure(bg=T["surface"], highlightbackground=T["line"])
                elif role == "chip_text":
                    w.configure(bg=T["surface"], fg=T["text"], font=self.fs(11))
                elif role == "chip_x":
                    w.configure(bg=T["surface"], fg=T["mut2"], activebackground=T["surface"],
                                activeforeground=T["err"], font=small)
                elif role == "detail":
                    w.configure(bg=T["surface"], fg=T["quote"], insertbackground=T["accent"],
                                font=self.fs(11))
                elif role == "log":
                    w.configure(bg=T["bg"], fg=T["mut"], insertbackground=T["accent"], font=small)
                elif role == "check":
                    w.configure(bg=T["bg"], fg=T["mut"], selectcolor=T["surface"],
                                activebackground=T["bg"], activeforeground=T["accent"], font=small)
                elif role == "check_s":
                    w.configure(bg=T["surface"], fg=T["mut"], selectcolor=T["bg"],
                                activebackground=T["surface"], activeforeground=T["accent"], font=small)
            except tk.TclError:
                pass

        for key, btn in self.mode_btns.items():
            on = (key == self.mode)
            btn.configure(bg=T["accent"] if on else T["surface"],
                          fg=T["on_accent"] if on else T["mut"],
                          activebackground=T["accent"] if on else T["surface"],
                          activeforeground=T["on_accent"] if on else T["accent"],
                          highlightbackground=T["line"], font=self.fs(9, kind="mono"))

        for key, btn in getattr(self, "score_btns", {}).items():
            on = (key == self.v_score.get())
            btn.configure(bg=T["accent"] if on else T["bg"],
                          fg=T["on_accent"] if on else T["mut2"],
                          activebackground=T["accent"] if on else T["surface"],
                          activeforeground=T["on_accent"] if on else T["accent"],
                          highlightbackground=T["line"], font=self.fs(9, kind="mono"))

        self.side_scroll.retheme(T["surface"])
        self.rail_scroll.retheme(T["bg"])

        st = ttk.Style(); st.theme_use("default")
        rowh = max(26, int(30 * self.scale))
        st.configure("V.Treeview", background=T["bg"], fieldbackground=T["bg"],
                     foreground=T["text"], rowheight=rowh, borderwidth=0, font=self.fs(11))
        st.configure("V.Treeview.Heading", background=T["surface"], foreground=T["mut"],
                     relief="flat", font=self.fs(9, kind="mono"))
        st.map("V.Treeview", background=[("selected", T["sel"])], foreground=[("selected", T["text"])])
        st.configure("V.Horizontal.TProgressbar", troughcolor=T["surface"],
                     background=T["accent"], borderwidth=0, thickness=max(4, int(5 * self.scale)))
        st.configure("V.TCombobox", fieldbackground=T["surface"], background=T["surface"],
                     foreground=T["text"], arrowcolor=T["accent"])
        st.configure("V.Vertical.TScrollbar", background=T["surface"], troughcolor=T["bg"],
                     bordercolor=T["line"], arrowcolor=T["mut"])
        self.tree.configure(style="V.Treeview")
        self.tree.tag_configure("high", foreground=T["accent"])
        self.tree.tag_configure("mid",  foreground=T["text"])
        self.tree.tag_configure("low",  foreground=T["mut2"])
        self.tree.tag_configure("inc",  background=T["sel"])
        self.tree.tag_configure("exc",  foreground=T["mut2"])
        for tag, col in (("accent", T["accent"]), ("ok", T["ok"]),
                         ("err", T["err"]), ("dim", T["mut2"])):
            self.log_box.tag_config(tag, foreground=col)
        self._render_chips()
        self._render_rail()
        self._render_detail()

    def _on_field_pick(self, _e=None):
        label = self.cb_field.get()
        key = "auto"
        for k, v in PROFILE_LABELS.items():
            if v == label:
                key = k
        self.v_field.set(key)
        if key != "auto":
            self.apply_profile(key, announce=True)
        else:
            self.log("Field will be detected from your question.", "dim")
        self._save_settings()

    def apply_profile(self, key, announce=False):
        """A field decides its rubric and which databases are worth querying."""
        self.profile = key
        src = PROFILE_SOURCES.get(key, PROFILE_SOURCES["general"])
        self.v_pubmed.set(src["pubmed"]); self.v_epmc.set(src["epmc"])
        self.v_arxiv.set(src["arxiv"]);   self.v_cross.set(src["crossref"])
        self.v_oa.set(src["openalex"])
        if announce:
            names = [n for n, on in (("PubMed", src["pubmed"]), ("Europe PMC", src["epmc"]),
                                     ("arXiv", src["arxiv"]), ("Crossref", src["crossref"]),
                                     ("OpenAlex", src["openalex"])) if on]
            self.log(f"{PROFILE_LABELS[key]} — searching {', '.join(names)}; "
                     f"rubric: {', '.join(l for _k, l, _d in rubric_for(key))}", "accent")

    def set_mode(self, mode):
        self.mode = mode
        self._apply_theme()
        self._render_rows()
        self._save_settings()

    def set_score_mode(self, key):
        self.v_score.set(key)
        self._apply_theme()
        self._save_settings()
        self.log({"ask":  "AI scoring: I'll ask after each search.",
                  "auto": "AI scoring: runs automatically after ranking.",
                  "off":  "AI scoring off — local ranking only."}[key], "dim")

    # ── build ─────────────────────────────────────────────────────────────────
    def _build(self):
        self.grid_rowconfigure(4, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self._build_header()
        self.reg(tk.Frame(self, height=1), "line").grid(row=1, column=0, sticky="ew")
        self._build_query()
        self._build_status()
        self._build_main()

    def _build_header(self):
        head = self.reg(tk.Frame(self), "surface")
        head.grid(row=0, column=0, sticky="ew")
        self.reg(tk.Label(head, text="VERONICA"), "brand").pack(side="left", padx=(18, 10), pady=9)
        self.reg(tk.Label(head, text=f"v{VERSION} · everything runs on this machine"),
                 "mono_s").pack(side="left")

        self.mode_btns = {}
        sw = self.reg(tk.Frame(head), "surface"); sw.pack(side="right", padx=(10, 18))
        for key, label in (("light", "LIGHT"), ("dark", "DARK")):
            b = tk.Button(sw, text=label, relief="flat", bd=0, padx=10, pady=3, cursor="hand2",
                          highlightthickness=1, command=lambda k=key: self.set_mode(k))
            b.pack(side="left")
            self.mode_btns[key] = b

        self.btn_full = self.reg(tk.Button(head, text="FULL SCREEN", relief="flat", bd=0,
                                           cursor="hand2", padx=10, highlightthickness=1,
                                           command=self.toggle_fullscreen), "ghost_s")
        self.btn_full.pack(side="right", padx=4)
        self.reg(tk.Button(head, text="SETTINGS", relief="flat", bd=0, cursor="hand2",
                           padx=10, highlightthickness=1,
                           command=self.open_settings), "ghost_s").pack(side="right", padx=4)
        self.lbl_model = self.reg(tk.Label(head, text="● checking ollama"), "mono_s")
        self.lbl_model.pack(side="right", padx=12)

    def _build_query(self):
        band = self.reg(tk.Frame(self), "bg")
        band.grid(row=2, column=0, sticky="ew")
        band.grid_columnconfigure(0, weight=1)

        # question
        qwrap = self.reg(tk.Frame(band), "bg")
        qwrap.grid(row=0, column=0, sticky="ew", padx=18, pady=(12, 0))
        qwrap.grid_columnconfigure(0, weight=1)
        self.reg(tk.Label(qwrap, text="YOUR RESEARCH QUESTION", anchor="w"),
                 "label").grid(row=0, column=0, sticky="ew")
        self.txt_q = self.reg(tk.Text(qwrap, height=2, relief="flat", bd=0, wrap="word",
                                      highlightthickness=1), "entry")
        self.txt_q.grid(row=1, column=0, sticky="ew", pady=(4, 0))
        self.txt_q.insert("1.0", self.settings.get("question", ""))
        self.txt_q.bind("<FocusOut>", lambda e: self.refresh_terms(auto=True))
        self.btn_run = self.reg(tk.Button(qwrap, text="  RUN REVIEW  ", relief="flat", bd=0,
                                          cursor="hand2", padx=18, command=self.start), "primary")
        self.btn_run.grid(row=1, column=1, sticky="ns", padx=(12, 0), pady=(4, 0))

        self.reg(tk.Label(qwrap, text="Ask it as a question — relevance is scored against it, "
                                      "word for word.", anchor="w"),
                 "hint").grid(row=2, column=0, sticky="ew", pady=(3, 0))

        # derived terms
        twrap = self.reg(tk.Frame(band), "bg")
        twrap.grid(row=1, column=0, sticky="ew", padx=18, pady=(12, 0))
        twrap.grid_columnconfigure(0, weight=1)
        hdr = self.reg(tk.Frame(twrap), "bg"); hdr.grid(row=0, column=0, sticky="ew")
        self.reg(tk.Label(hdr, text="SEARCHING FOR", anchor="w"), "label").pack(side="left")
        self.reg(tk.Button(hdr, text="re-read question", relief="flat", bd=0, cursor="hand2",
                           padx=8, highlightthickness=1,
                           command=lambda: self.refresh_terms(force=True)),
                 "ghost").pack(side="right")
        self.reg(tk.Button(hdr, text="show query", relief="flat", bd=0, cursor="hand2",
                           padx=8, highlightthickness=1,
                           command=self.toggle_query), "ghost").pack(side="right", padx=6)

        # field profile — decides the rubric and which sources are worth asking
        fwrap = self.reg(tk.Frame(hdr), "bg"); fwrap.pack(side="right", padx=(0, 16))
        self.reg(tk.Label(fwrap, text="FIELD", anchor="e"), "label").pack(side="left", padx=(0, 7))
        self.cb_field = ttk.Combobox(fwrap, width=26, state="readonly", style="V.TCombobox",
                                     values=["Detect from my question"] +
                                            [PROFILE_LABELS[k] for k in
                                             ("life", "physical", "computing", "social", "general")])
        self.cb_field.pack(side="left")
        self.cb_field.set("Detect from my question" if self.v_field.get() == "auto"
                          else PROFILE_LABELS.get(self.v_field.get(), "Detect from my question"))
        self.cb_field.bind("<<ComboboxSelected>>", self._on_field_pick)

        # AI scoring switch — out in the open, because it is the expensive choice
        self.score_btns = {}
        sw = self.reg(tk.Frame(hdr), "bg"); sw.pack(side="right", padx=(0, 16))
        self.reg(tk.Label(sw, text="AI SCORING", anchor="e"), "label").pack(side="left", padx=(0, 7))
        for key, label in (("ask", "ASK ME"), ("auto", "AUTOMATIC"), ("off", "OFF")):
            b = tk.Button(sw, text=label, relief="flat", bd=0, padx=9, pady=2, cursor="hand2",
                          highlightthickness=1, command=lambda k=key: self.set_score_mode(k))
            b.pack(side="left")
            self.score_btns[key] = b
        self.chip_box = self.reg(tk.Frame(twrap), "bg")
        self.chip_box.grid(row=1, column=0, sticky="ew", pady=(6, 0))
        self.reg(tk.Label(twrap, anchor="w", justify="left",
                          text="These are the ideas that must ALL appear in a paper. Veronica "
                               "reads them out of your question; drop one to widen the search, "
                               "add one to narrow it."),
                 "hint").grid(row=2, column=0, sticky="ew", pady=(5, 0))

        # collapsible generated query
        self.query_wrap = self.reg(tk.Frame(band), "bg")
        self.query_wrap.grid_columnconfigure(1, weight=1)
        self.reg(tk.Label(self.query_wrap, text="QUERY SENT →", anchor="nw"),
                 "label").grid(row=0, column=0, sticky="nw", padx=(0, 8))
        self.txt_query = self.reg(tk.Text(self.query_wrap, height=3, relief="flat", bd=0,
                                          wrap="word", highlightthickness=1), "query")
        self.txt_query.grid(row=0, column=1, sticky="ew")
        self.query_open = False

        self.reg(tk.Frame(band, height=14), "bg").grid(row=3, column=0)

    def toggle_query(self):
        self.query_open = not self.query_open
        if self.query_open:
            self.preview_query()
            self.query_wrap.grid(row=2, column=0, sticky="ew", padx=18, pady=(12, 0))
        else:
            self.query_wrap.grid_remove()

    def _build_status(self):
        strip = self.reg(tk.Frame(self), "surface")
        strip.grid(row=3, column=0, sticky="ew")
        strip.grid_columnconfigure(1, weight=1)
        self.reg(tk.Frame(strip, height=1), "line").grid(row=0, column=0, columnspan=3, sticky="ew")
        self.lbl_status = self.reg(tk.Label(strip, text="idle", anchor="w"), "mono_s")
        self.lbl_status.grid(row=1, column=0, sticky="w", padx=16, pady=6)
        self.pbar = ttk.Progressbar(strip, mode="determinate", maximum=100,
                                    style="V.Horizontal.TProgressbar")
        self.pbar.grid(row=1, column=1, sticky="ew", padx=8)
        self.lbl_elapsed = self.reg(tk.Label(strip, text=""), "mono_s")
        self.lbl_elapsed.grid(row=1, column=2, sticky="e", padx=16)

    def _build_main(self):
        outer = tk.PanedWindow(self, orient="horizontal", sashwidth=6, bd=0,
                               bg=self.T["line"], sashrelief="flat")
        self.reg(outer, "line")
        outer.grid(row=4, column=0, sticky="nsew")

        # left rail
        rail = self.reg(tk.Frame(outer), "bg")
        self.rail_scroll = ScrollFrame(rail, self.T["bg"])
        self.rail_scroll.pack(fill="both", expand=True)
        outer.add(rail, minsize=170, width=int(200 * self.scale), stretch="never")

        # centre
        centre = tk.PanedWindow(outer, orient="vertical", sashwidth=6, bd=0,
                                bg=self.T["line"], sashrelief="flat")
        self.reg(centre, "line")
        table = self.reg(tk.Frame(centre), "bg")
        self.ask_bar = self.reg(tk.Frame(table), "surface")
        cols = ("score", "title", "journal", "year", "cites", "source", "pdf", "screen")
        self.tree = ttk.Treeview(table, columns=cols, show="headings", selectmode="browse")
        widths = {"score": 84, "title": 480, "journal": 150, "year": 56,
                  "cites": 62, "source": 132, "pdf": 48, "screen": 82}
        for col, hdr, anchor in (("score", "SCORE /12", "center"), ("title", "TITLE", "w"),
                                 ("journal", "JOURNAL", "w"), ("year", "YEAR", "center"),
                                 ("cites", "CITED", "center"), ("source", "SOURCE", "w"),
                                 ("pdf", "PDF", "center"), ("screen", "SCREEN", "center")):
            self.tree.heading(col, text=hdr, command=lambda c=col: self.sort_by(c))
            self.tree.column(col, width=int(widths[col] * self.scale),
                             minwidth=40, anchor=anchor, stretch=(col == "title"))
        vsb = ttk.Scrollbar(table, orient="vertical", command=self.tree.yview,
                            style="V.Vertical.TScrollbar")
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self.on_select)
        self.tree.bind("<Double-1>", lambda e: self.open_pdf())
        self.bind_all("<Key>", self._hotkey)
        self.empty = self.reg(tk.Label(table, text="", anchor="center", justify="center"), "hint")
        centre.add(table, minsize=200, stretch="always")

        logwrap = self.reg(tk.Frame(centre), "bg")
        bar = self.reg(tk.Frame(logwrap), "bg"); bar.pack(fill="x")
        self.reg(tk.Label(bar, text="ACTIVITY", anchor="w"), "label").pack(side="left", padx=14, pady=4)
        self.log_box = self.reg(tk.Text(logwrap, height=6, relief="flat", bd=0, wrap="word",
                                        state="disabled"), "log")
        self.log_box.pack(fill="both", expand=True, padx=10, pady=(0, 8))
        centre.add(logwrap, minsize=70, height=int(140 * self.scale), stretch="never")
        outer.add(centre, minsize=420, stretch="always")

        # right detail
        side = self.reg(tk.Frame(outer), "surface")
        self.side_scroll = ScrollFrame(side, self.T["surface"])
        self.side_scroll.pack(fill="both", expand=True)
        inner = self.side_scroll.inner

        self.reg(tk.Label(inner, text="WHY THIS SCORE", anchor="w"),
                 "label_s").pack(fill="x", padx=16, pady=(14, 6))
        self.lbl_title = self.reg(tk.Label(inner, text="Select a paper", anchor="w",
                                           justify="left", wraplength=int(320 * self.scale)),
                                  "text_s")
        self.lbl_title.pack(fill="x", padx=16)
        self.lbl_meta = self.reg(tk.Label(inner, text="", anchor="w", justify="left",
                                          wraplength=int(320 * self.scale)), "mono_s")
        self.lbl_meta.pack(fill="x", padx=16, pady=(4, 10))
        self.rubric_box = self.reg(tk.Frame(inner), "surface")
        self.rubric_box.pack(fill="x", padx=16)
        self.reg(tk.Label(inner, text="SUMMARY", anchor="w"),
                 "label_s").pack(fill="x", padx=16, pady=(14, 4))
        self.txt_detail = self.reg(tk.Text(inner, height=12, relief="flat", bd=0, wrap="word",
                                           state="disabled"), "detail")
        self.txt_detail.pack(fill="x", padx=16)
        act = self.reg(tk.Frame(inner), "surface"); act.pack(fill="x", padx=14, pady=(14, 4))
        self.reg(tk.Button(act, text="OPEN PDF", relief="flat", bd=0, cursor="hand2",
                           highlightthickness=1, padx=8, pady=6,
                           command=self.open_pdf), "primary").pack(fill="x")
        act2 = self.reg(tk.Frame(inner), "surface"); act2.pack(fill="x", padx=14, pady=(6, 16))
        for text, cmd in (("SCORE", self.score_one),
                          ("INCLUDE", lambda: self.screen("include")),
                          ("EXCLUDE", lambda: self.screen("exclude"))):
            self.reg(tk.Button(act2, text=text, relief="flat", bd=0, cursor="hand2",
                               highlightthickness=1, padx=4, pady=5, command=cmd),
                     "ghost_s").pack(side="left", expand=True, fill="x", padx=2)
        outer.add(side, minsize=280, width=int(360 * self.scale), stretch="never")

    # ── search terms (chips) ──────────────────────────────────────────────────
    def refresh_terms(self, auto=False, force=False):
        q = self.txt_q.get("1.0", "end").strip()
        if not q:
            return
        if auto and self.terms and not force:
            return
        derived = derive_terms(q)
        if derived and (force or not self.terms):
            self.terms = derived
            self._render_chips()
            if self.query_open:
                self.preview_query()

    def add_term(self):
        T = self.T
        win = tk.Toplevel(self); win.title("Add a term"); win.configure(bg=T["bg"])
        win.transient(self); win.resizable(False, False)
        tk.Label(win, text="Papers must mention this idea:", bg=T["bg"], fg=T["mut"],
                 font=self.fs(10, kind="mono")).pack(padx=18, pady=(16, 6))
        var = tk.StringVar()
        e = tk.Entry(win, textvariable=var, width=34, relief="flat", bd=0,
                     highlightthickness=1, bg=T["surface"], fg=T["text"],
                     insertbackground=T["accent"], highlightbackground=T["line"],
                     font=self.fs(12))
        e.pack(padx=18, ipady=5); e.focus_set()

        def ok(_e=None):
            v = var.get().strip()
            if v and v.lower() not in [t.lower() for t in self.terms]:
                self.terms.append(v)
                self._render_chips()
                if self.query_open:
                    self.preview_query()
            win.destroy()

        e.bind("<Return>", ok)
        tk.Button(win, text="ADD", relief="flat", bd=0, cursor="hand2", padx=14, pady=5,
                  bg=T["accent"], fg=T["on_accent"], activebackground=T["accent_dim"],
                  font=self.fs(11, "bold", "head"), command=ok).pack(pady=14)

    def drop_term(self, term):
        self.terms = [t for t in self.terms if t != term]
        self._render_chips()
        if self.query_open:
            self.preview_query()

    def _render_chips(self):
        if not hasattr(self, "chip_box"):
            return
        for w in self.chip_box.winfo_children():
            w.destroy()
        T = self.T
        if not self.terms:
            self.reg(tk.Label(self.chip_box, text="— type a question above, then "
                                                  "\"re-read question\" —", anchor="w"),
                     "hint").pack(side="left")
        for t in self.terms:
            chip = tk.Frame(self.chip_box, bg=T["surface"], highlightthickness=1,
                            highlightbackground=T["line"])
            chip.pack(side="left", padx=(0, 8), pady=2)
            tk.Label(chip, text=t, bg=T["surface"], fg=T["text"],
                     font=self.fs(11)).pack(side="left", padx=(9, 4), pady=3)
            n = len(concept_variants(t))
            if n > 1:
                tk.Label(chip, text=f"+{n-1}", bg=T["surface"], fg=T["accent"],
                         font=self.fs(9, kind="mono")).pack(side="left", padx=(0, 2))
            tk.Button(chip, text="×", relief="flat", bd=0, cursor="hand2", padx=6,
                      bg=T["surface"], fg=T["mut2"], activebackground=T["surface"],
                      activeforeground=T["err"], font=self.fs(11, kind="mono"),
                      command=lambda t=t: self.drop_term(t)).pack(side="left")
        tk.Button(self.chip_box, text="+ add", relief="flat", bd=0, cursor="hand2", padx=10,
                  pady=3, bg=T["bg"], fg=T["accent"], activebackground=T["surface"],
                  activeforeground=T["accent"], highlightthickness=1,
                  highlightbackground=T["line"], font=self.fs(10, kind="mono"),
                  command=self.add_term).pack(side="left", pady=2)

    # ── settings dialog ───────────────────────────────────────────────────────
    def open_settings(self):
        T = self.T
        win = tk.Toplevel(self); win.title("Settings"); win.configure(bg=T["bg"])
        win.transient(self)
        win.geometry(f"{int(430*self.scale)}x{int(430*self.scale)}")
        pad = dict(padx=20)

        def section(text):
            tk.Label(win, text=text, anchor="w", bg=T["bg"], fg=T["mut2"],
                     font=self.fs(9, kind="mono")).pack(fill="x", pady=(16, 6), **pad)

        def row(label, var, values, hint=""):
            r = tk.Frame(win, bg=T["bg"]); r.pack(fill="x", **pad)
            tk.Label(r, text=label, anchor="w", bg=T["bg"], fg=T["text"],
                     font=self.fs(11)).pack(side="left")
            if values:
                ttk.Combobox(r, textvariable=var, values=values, width=8,
                             state="readonly", style="V.TCombobox").pack(side="right")
            else:
                tk.Entry(r, textvariable=var, width=10, relief="flat", bd=0, justify="right",
                         highlightthickness=1, bg=T["surface"], fg=T["text"],
                         insertbackground=T["accent"], highlightbackground=T["line"],
                         font=self.fs(11)).pack(side="right", ipady=3)
            if hint:
                tk.Label(win, text=hint, anchor="w", justify="left", bg=T["bg"], fg=T["mut2"],
                         wraplength=int(370 * self.scale),
                         font=self.fs(9)).pack(fill="x", **pad)

        section("HOW WIDE TO SEARCH")
        row("Retrieve per source", self.v_retrieve, ["50", "100", "200", "300", "500"],
            "How many records to pull before ranking. Wider costs seconds, not minutes.")
        row("Published from", self.v_year, None, "Blank means no year limit.")

        section("MODEL")
        row("Ollama model", self.v_model, None, "Used for drafting the review.")
        row("Parallel requests", self.v_workers, ["1", "2", "4", "6", "8"],
            "4 is safe on 16 GB. Raise it if scoring is slow and RAM allows.")

        section("AI SCORING — THE SLOW PART")
        tk.Label(win, text="Local ranking is instant and always runs. The model adds the "
                          "0–12 rubric and evidence quotes at roughly 15–30 s per paper. "
                          "The Ask / Auto / Off switch is in the main window, next to "
                          "SEARCHING FOR.",
                 anchor="w", justify="left", bg=T["bg"], fg=T["mut2"],
                 wraplength=int(370 * self.scale), font=self.fs(9)).pack(fill="x", **pad)
        row("Score the top", self.v_keep, ["10", "20", "40", "60", "100"],
            "How many papers a batch covers.")
        row("Scoring model", self.v_smodel, None,
            "Scoring is triage, so a small model is the right tool — llama3.2:3b is roughly "
            "three times faster than 8b and rarely changes which papers rise. Drafting uses "
            "the main model below.")
        row("Papers per call", self.v_batch, ["1", "3", "5", "8"],
            "Scoring several papers in one request pays the prompt and warm-up once. "
            "5 is the sweet spot; drop to 1 if a small model starts confusing papers.")
        tk.Checkbutton(win, variable=self.v_fast,
                       text="Fast mode — shorter context, about twice as quick",
                       anchor="w", relief="flat", bd=0, highlightthickness=0, bg=T["bg"],
                       fg=T["mut"], selectcolor=T["surface"], activebackground=T["bg"],
                       activeforeground=T["accent"],
                       font=self.fs(10)).pack(fill="x", padx=18, pady=(4, 0))

        section("SOURCES AND FILES")
        tk.Checkbutton(win, variable=self.v_pubmed,
                       text="PubMed — biomedical literature only",
                       anchor="w", relief="flat", bd=0, highlightthickness=0, bg=T["bg"],
                       fg=T["mut"], selectcolor=T["surface"], activebackground=T["bg"],
                       activeforeground=T["accent"],
                       font=self.fs(10)).pack(fill="x", padx=18, pady=1)
        for var, text in ((self.v_epmc, "Europe PMC — preprints and open-access full text"),
                          (self.v_oa, "OpenAlex — every field, citation counts"),
                          (self.v_arxiv, "arXiv — physics, maths, CS, quant bio"),
                          (self.v_cross, "Crossref — every discipline with a DOI"),
                          (self.v_pdf, "Download every shortlisted PDF up front (slower)"),
                          (self.v_nofl, "Keep papers with no full-text link")):
            tk.Checkbutton(win, variable=var, text=text, anchor="w", relief="flat", bd=0,
                           highlightthickness=0, bg=T["bg"], fg=T["mut"],
                           selectcolor=T["surface"], activebackground=T["bg"],
                           activeforeground=T["accent"],
                           font=self.fs(10)).pack(fill="x", padx=18, pady=1)

        btns = tk.Frame(win, bg=T["bg"]); btns.pack(fill="x", pady=18, **pad)
        tk.Button(btns, text="CHECK MODEL", relief="flat", bd=0, cursor="hand2", padx=12, pady=5,
                  bg=T["bg"], fg=T["text"], highlightthickness=1, highlightbackground=T["line"],
                  activeforeground=T["accent"], font=self.fs(10, kind="mono"),
                  command=self.check_model).pack(side="left")
        tk.Button(btns, text="OUTPUT FOLDER", relief="flat", bd=0, cursor="hand2", padx=12, pady=5,
                  bg=T["bg"], fg=T["text"], highlightthickness=1, highlightbackground=T["line"],
                  activeforeground=T["accent"], font=self.fs(10, kind="mono"),
                  command=self.open_folder).pack(side="left", padx=6)
        tk.Button(btns, text="DONE", relief="flat", bd=0, cursor="hand2", padx=16, pady=5,
                  bg=T["accent"], fg=T["on_accent"], activebackground=T["accent_dim"],
                  font=self.fs(11, "bold", "head"),
                  command=lambda: (self._save_settings(), win.destroy())).pack(side="right")

    # ── messaging ─────────────────────────────────────────────────────────────
    def post(self, kind, **kw):
        self.msgq.put((kind, kw))

    def log(self, msg, tag=""):
        self.post("log", msg=msg, tag=tag)

    def _drain(self):
        try:
            while True:
                kind, kw = self.msgq.get_nowait()
                if kind == "log":
                    self.log_box.configure(state="normal")
                    ts = datetime.datetime.now().strftime("%H:%M:%S")
                    self.log_box.insert("end", f"[{ts}] ", "dim")
                    self.log_box.insert("end", kw["msg"] + "\n", kw.get("tag") or "")
                    self.log_box.see("end")
                    self.log_box.configure(state="disabled")
                elif kind == "status":
                    self.lbl_status.configure(text=kw["text"])
                    if "pct" in kw:
                        self.pbar.configure(value=kw["pct"])
                    if "elapsed" in kw:
                        self.lbl_elapsed.configure(text=kw["elapsed"])
                elif kind == "rows":     self._render_rows()
                elif kind == "themes":   self._render_rail()
                elif kind == "detail":   self._render_detail()
                elif kind == "ask":      self._show_ask(kw["n"])
                elif kind == "scoring":
                    self.btn_run.configure(text="  STOP SCORING  " if kw["on"]
                                                else "  RUN REVIEW  ", state="normal")
                elif kind == "query":
                    self.txt_query.delete("1.0", "end")
                    self.txt_query.insert("end", kw["text"])
                elif kind == "done":
                    self.running = False
                    self.btn_run.configure(text="  RUN REVIEW  ", state="normal")
                elif kind == "synthesis":
                    self._show_synthesis(kw["text"], kw["path"])
        except queue.Empty:
            pass
        self.after(120, self._drain)

    # ── rendering ─────────────────────────────────────────────────────────────
    @staticmethod
    def _tier(total):
        if total is None:
            return "low"
        return "high" if total >= 9 else "mid" if total >= 5 else "low"

    def _render_rows(self):
        want = self.theme_filter.get()
        self.view = [p for p in self.papers if want == "All themes" or p.get("theme") == want]
        self.tree.delete(*self.tree.get_children())
        for p in self.view:
            s = p.get("scored") or {}
            total = s.get("total")
            tags = [self._tier(total)]
            d = self.decisions.get(p["key"], "")
            if d == "include": tags.append("inc")
            if d == "exclude": tags.append("exc")
            self.tree.insert("", "end", iid=p["key"], tags=tuple(tags), values=(
                total if total is not None else
                ("✓" * int(round((p.get("match") or {}).get("coverage", 0) * 3)) or "·"),
                p.get("title", "")[:150],
                (p.get("journal", "") or "")[:34],
                p.get("year", ""),
                p.get("cited_by") if p.get("cited_by") is not None else "·",
                p.get("source", ""),
                "✓" if p.get("pdf_path") else ("link" if p.get("has_fulltext") else "·"),
                {"include": "IN", "exclude": "OUT"}.get(d, ""),
            ))
        if self.view:
            self.empty.place_forget()
        else:
            self.empty.configure(text=self._empty_text())
            self.empty.place(relx=0.5, rely=0.45, anchor="center")
        self._render_rail()

    def _empty_text(self):
        if self.running and not self.scoring:
            return ("Searching…\n\n"
                    "Sources are queried one at a time, then deduplicated and ranked "
                    "locally.\nResults appear here the moment ranking finishes — "
                    "usually a few seconds.")
        if self.papers:
            return ("Nothing in this theme.\n\nPick “All themes” in the left rail "
                    "to see everything again.")
        return ("No papers yet.\n\n"
                "1 · Type a research question above\n"
                "2 · Check the terms Veronica pulled out of it\n"
                "3 · Press RUN REVIEW\n\n"
                "Ranking is local and takes seconds. AI scoring is optional — "
                "the switch is next to SEARCHING FOR.")

    def _render_rail(self):
        if not hasattr(self, "rail_scroll"):
            return
        box = self.rail_scroll.inner
        for w in box.winfo_children():
            w.destroy()
        T = self.T
        mono, small = self.fs(10, kind="mono"), self.fs(9, kind="mono")

        tk.Label(box, text="THEMES", anchor="w", bg=T["bg"], fg=T["mut2"],
                 font=small).pack(fill="x", padx=14, pady=(14, 6))
        opts = ["All themes"] + self.themes
        for name in opts:
            n = len(self.papers) if name == "All themes" else \
                sum(1 for p in self.papers if p.get("theme") == name)
            b = tk.Radiobutton(box, text=f"{name}  ({n})", value=name, variable=self.theme_filter,
                               anchor="w", relief="flat", bd=0, highlightthickness=0,
                               indicatoron=False, command=self._render_rows, cursor="hand2",
                               padx=8, pady=3, bg=T["bg"], fg=T["text"], selectcolor=T["sel"],
                               activebackground=T["surface"], activeforeground=T["accent"],
                               font=small)
            b.pack(fill="x", padx=10, pady=1)

        inc = sum(1 for v in self.decisions.values() if v == "include")
        exc = sum(1 for v in self.decisions.values() if v == "exclude")
        scored = sum(1 for p in self.papers if p.get("scored"))
        tk.Label(box, text="SCREENING", anchor="w", bg=T["bg"], fg=T["mut2"],
                 font=small).pack(fill="x", padx=14, pady=(18, 6))
        for label, val in (("retrieved", len(self.papers)), ("scored", scored),
                           ("included", inc), ("excluded", exc), ("shown", len(self.view))):
            r = tk.Frame(box, bg=T["bg"]); r.pack(fill="x", padx=14)
            tk.Label(r, text=label, anchor="w", bg=T["bg"], fg=T["mut"],
                     font=small).pack(side="left")
            tk.Label(r, text=str(val), anchor="e", bg=T["bg"],
                     fg=T["accent"] if label == "included" else T["text"],
                     font=mono).pack(side="right")

        tk.Label(box, text="KEYS", anchor="w", bg=T["bg"], fg=T["mut2"],
                 font=small).pack(fill="x", padx=14, pady=(18, 6))
        tk.Label(box, text="I  include\nX  exclude\nU  undo\nJ / K  move\n"
                          "Enter  open PDF\nF11  full screen",
                 anchor="w", justify="left", bg=T["bg"], fg=T["mut"],
                 font=small).pack(fill="x", padx=14)

        tk.Button(box, text="draft review", relief="flat", bd=0, cursor="hand2", padx=10, pady=5,
                  bg=T["bg"], fg=T["text"], highlightthickness=1, highlightbackground=T["line"],
                  activeforeground=T["accent"], font=small,
                  command=self.do_synthesis).pack(fill="x", padx=14, pady=(18, 4))
        tk.Button(box, text="overview charts", relief="flat", bd=0, cursor="hand2",
                  padx=10, pady=5, bg=T["bg"], fg=T["text"], highlightthickness=1,
                  highlightbackground=T["line"], activeforeground=T["accent"], font=small,
                  command=self.open_overview).pack(fill="x", padx=14, pady=4)
        tk.Button(box, text="export citations", relief="flat", bd=0, cursor="hand2",
                  padx=10, pady=5, bg=T["bg"], fg=T["text"], highlightthickness=1,
                  highlightbackground=T["line"], activeforeground=T["accent"], font=small,
                  command=self.export_citations).pack(fill="x", padx=14, pady=(4, 18))

    def _render_detail(self):
        if not hasattr(self, "rubric_box"):
            return
        for w in self.rubric_box.winfo_children():
            w.destroy()
        p, T = self.selected, self.T
        mono, body = self.fs(9, kind="mono"), self.fs(10)
        if not p:
            self.lbl_title.configure(text="Select a paper")
            self.lbl_meta.configure(text="The four sub-scores and the sentence each one "
                                         "was taken from will show here.")
            self._set_detail("")
            return

        self.lbl_title.configure(text=p.get("title", ""))
        bits = [fmt_authors(p.get("authors"), 3), p.get("journal", ""), p.get("year", "")]
        if p.get("cited_by") is not None:
            bits.append(f"{p['cited_by']} cites")
        if p.get("prerank") is not None:
            bits.append(f"pre-rank {p['prerank']}")
        self.lbl_meta.configure(text=" · ".join([b for b in bits if b]))

        # local match explanation — always available, no model involved
        m = p.get("match") or {}
        if m:
            hits = m.get("hits") or []
            miss = m.get("missing") or []
            tk.Label(self.rubric_box, text="WHY IT RANKED HERE", anchor="w",
                     bg=T["surface"], fg=T["mut2"], font=mono).pack(fill="x", pady=(0, 4))
            for name, where in hits:
                r = tk.Frame(self.rubric_box, bg=T["surface"]); r.pack(fill="x")
                tk.Label(r, text="✓", bg=T["surface"], fg=T["accent"],
                         font=mono).pack(side="left", padx=(0, 6))
                tk.Label(r, text=name, anchor="w", bg=T["surface"], fg=T["text"],
                         font=body).pack(side="left")
                tk.Label(r, text=f"in {where}", anchor="e", bg=T["surface"], fg=T["mut2"],
                         font=mono).pack(side="right")
            for name in miss:
                r = tk.Frame(self.rubric_box, bg=T["surface"]); r.pack(fill="x")
                tk.Label(r, text="–", bg=T["surface"], fg=T["mut2"],
                         font=mono).pack(side="left", padx=(0, 6))
                tk.Label(r, text=name, anchor="w", bg=T["surface"], fg=T["mut2"],
                         font=body).pack(side="left")
                tk.Label(r, text="not found", anchor="e", bg=T["surface"], fg=T["mut2"],
                         font=mono).pack(side="right")
            tk.Frame(self.rubric_box, bg=T["line"], height=1).pack(fill="x", pady=10)

        s = p.get("scored") or {}
        crit = s.get("criteria") or []
        if isinstance(crit, dict):      # scores from an older run
            crit = [{"label": k.upper(), "score": v.get("score", 0),
                     "evidence": v.get("evidence", "")} for k, v in crit.items()]
        if not crit:
            tk.Label(self.rubric_box, text="Not scored — outside the top slice, or "
                                           "the model wasn't running. Press S to score it.",
                     anchor="w", justify="left", wraplength=int(320 * self.scale),
                     bg=T["surface"], fg=T["mut"], font=mono).pack(fill="x")
        else:
            tk.Label(self.rubric_box,
                     text=f"{s.get('total', 0)} / 12   {s.get('paper_type','')}",
                     anchor="w", bg=T["surface"], fg=T["accent"],
                     font=self.fs(15, "bold", "head")).pack(fill="x", pady=(0, 2))
            prof = s.get("profile")
            if prof:
                tk.Label(self.rubric_box, text=PROFILE_LABELS.get(prof, prof) + " rubric",
                         anchor="w", bg=T["surface"], fg=T["mut2"],
                         font=mono).pack(fill="x", pady=(0, 6))
        for node in crit:
            val = node.get("score", 0)
            row = tk.Frame(self.rubric_box, bg=T["surface"]); row.pack(fill="x", pady=(5, 0))
            tk.Label(row, text=node.get("label", ""), width=17, anchor="w", bg=T["surface"],
                     fg=T["mut"], font=mono).pack(side="left")
            bars = tk.Frame(row, bg=T["surface"]); bars.pack(side="left")
            for i in range(3):
                tk.Frame(bars, width=int(16 * self.scale), height=5,
                         bg=T["accent"] if i < val else T["bar_off"]).pack(side="left", padx=1)
            tk.Label(row, text=f" {val}/3", bg=T["surface"], fg=T["accent"],
                     font=mono).pack(side="left")
            ev = (node.get("evidence") or "").strip()
            if ev:
                tk.Label(self.rubric_box, text=f"“{ev}”", anchor="w", justify="left",
                         wraplength=int(310 * self.scale), bg=T["surface"], fg=T["quote"],
                         font=body).pack(fill="x", padx=(6, 0), pady=(2, 0))

        text = s.get("summary", "") or (p.get("abstract", "")[:700] or "No abstract on record.")
        extra = []
        if p.get("doi"):  extra.append(f"DOI: https://doi.org/{p['doi']}")
        if p.get("pmid"): extra.append(f"PubMed: https://pubmed.ncbi.nlm.nih.gov/{p['pmid']}/")
        if p.get("pdf_path"): extra.append(f"PDF: {p['pdf_path']}")
        elif not p.get("has_fulltext"): extra.append("No full-text link — abstract only.")
        self._set_detail(text + ("\n\n" + "\n".join(extra) if extra else ""))

    def _set_detail(self, text):
        self.txt_detail.configure(state="normal")
        self.txt_detail.delete("1.0", "end")
        self.txt_detail.insert("end", text)
        self.txt_detail.configure(state="disabled")

    # ── interactions ──────────────────────────────────────────────────────────
    def on_select(self, _e=None):
        sel = self.tree.selection()
        if not sel:
            return
        self.selected = next((p for p in self.papers if p["key"] == sel[0]), None)
        self._render_detail()

    def _hotkey(self, event):
        if isinstance(self.focus_get(), (tk.Text, tk.Entry, ttk.Combobox, ttk.Entry)):
            return
        k = (event.keysym or "").lower()
        if   k == "i": self.screen("include")
        elif k == "x": self.screen("exclude")
        elif k == "u": self.screen("")
        elif k in ("j", "down"): self._step(1)
        elif k in ("k", "up"):   self._step(-1)
        elif k == "s":           self.score_one()
        elif k == "return":      self.open_pdf()

    def _step(self, delta):
        if not self.view:
            return
        keys = [p["key"] for p in self.view]
        cur = self.selected["key"] if self.selected else None
        i = keys.index(cur) + delta if cur in keys else 0
        i = max(0, min(len(keys) - 1, i))
        self.tree.selection_set(keys[i])
        self.tree.see(keys[i])

    def screen(self, decision):
        if not self.selected:
            return
        key = self.selected["key"]
        if decision:
            self.decisions[key] = decision
        else:
            self.decisions.pop(key, None)
        self._render_rows()
        if key in [p["key"] for p in self.view]:
            self.tree.selection_set(key)

    def sort_by(self, col):
        keymap = {
            "score":   lambda p: (p.get("scored") or {}).get("total", -1),
            "title":   lambda p: p.get("title", "").lower(),
            "journal": lambda p: (p.get("journal") or "").lower(),
            "year":    lambda p: p.get("year", ""),
            "cites":   lambda p: p.get("cited_by") or -1,
            "source":  lambda p: p.get("source", ""),
            "pdf":     lambda p: bool(p.get("pdf_path")),
            "screen":  lambda p: self.decisions.get(p["key"], ""),
        }
        rev = col in ("score", "year", "cites", "pdf")
        self.papers.sort(key=keymap.get(col, keymap["score"]), reverse=rev)
        self._render_rows()

    def open_pdf(self):
        p = self.selected
        if not p:
            return
        if p.get("pdf_path") and Path(p["pdf_path"]).exists():
            self._open_path(p["pdf_path"])
            return
        if p.get("pmcid") and self.folder:
            # fetch on demand rather than downloading the whole shortlist up front
            def job():
                self.log(f"Fetching PDF for {p['title'][:50]}…", "dim")
                path = download_pdf(p, self.folder, self.log)
                if path:
                    p["pdf_path"] = path
                    self.post("rows")
                    self.after(0, lambda: self._open_path(path))
                elif p.get("fulltext_url"):
                    self.log("No open PDF — opening the publisher page.", "dim")
                    webbrowser.open(p["fulltext_url"])
            threading.Thread(target=job, daemon=True).start()
            return
        if p.get("fulltext_url"):
            webbrowser.open(p["fulltext_url"])
        else:
            messagebox.showinfo("Veronica", "No full-text link for this paper.")

    def open_folder(self):
        target = self.folder or BASE_DIR
        target.mkdir(parents=True, exist_ok=True)
        self._open_path(str(target))

    @staticmethod
    def _open_path(path):
        if sys.platform == "win32":    os.startfile(path)          # noqa
        elif sys.platform == "darwin": os.system(f'open "{path}"')
        else:                          os.system(f'xdg-open "{path}"')

    def preview_query(self):
        if not self.terms:
            self.refresh_terms(force=True)
        q = build_pubmed_query(self.terms, self.v_year.get())
        self.txt_query.delete("1.0", "end")
        self.txt_query.insert("end", q or "— no search terms yet —")
        return q

    def check_model(self):
        model = self.v_model.get().strip() or DEFAULT_MODEL
        ok, msg = check_ollama(model)
        if ok:
            self.lbl_model.configure(text=f"● ollama · {msg}", fg=self.T["ok"])
        else:
            self.lbl_model.configure(text="● ollama offline", fg=self.T["err"])
            self.log(msg, "err")

    # ── run ───────────────────────────────────────────────────────────────────
    def start(self):
        if self.running:
            self.running = False
            self.log("Stopping…", "err")
            return
        question = self.txt_q.get("1.0", "end").strip()
        if not question:
            messagebox.showwarning("Veronica", "Type a research question first — "
                                               "every score is measured against it.")
            return
        if not self.terms:
            self.refresh_terms(force=True)
        if not self.terms:
            messagebox.showwarning("Veronica", "I couldn't pull search terms out of that "
                                               "question. Add one with “+ add”.")
            return
        self._save_settings()
        self.running = True
        self.papers, self.decisions, self.themes, self.selected = [], {}, [], None
        self.theme_filter.set("All themes")
        self._render_rows(); self._render_detail()
        self.btn_run.configure(text="  STOP  ")
        threading.Thread(target=self._worker, args=(question, list(self.terms)),
                         daemon=True).start()

    def _status(self, text, pct=None, t0=None):
        kw = {"text": text}
        if pct is not None:
            kw["pct"] = pct
        if t0 is not None:
            secs = int(time.time() - t0)
            kw["elapsed"] = f"{secs // 60} m {secs % 60:02d} s" if secs >= 60 else f"{secs} s"
        self.post("status", **kw)

    def _worker(self, question, concepts):
        t0 = time.time()
        try:
            retrieve = int(self.v_retrieve.get() or 200)
            keep     = int(self.v_keep.get() or 40)
            workers  = max(1, int(self.v_workers.get() or 4))
            year     = self.v_year.get().strip()
            model    = self.v_model.get().strip() or DEFAULT_MODEL

            pubq = build_pubmed_query(concepts, year)
            self._status("building query", 2)
            self.post("query", text=pubq)

            # field decides the rubric and which databases get asked
            profile = self.v_field.get()
            if profile == "auto":
                profile = detect_field(question, concepts)
                self.after(0, lambda p=profile: self.apply_profile(p, announce=True))
            else:
                self.profile = profile
            self.log(f"Searching for: {' AND '.join(concepts)}", "accent")

            found = []
            if self.v_pubmed.get():
                self._status("searching PubMed", 6, t0)
                found += search_pubmed(pubq, retrieve, self.log)
            if self.v_epmc.get() and self.running:
                self._status("searching Europe PMC", 12, t0)
                found += search_europepmc(build_epmc_query(concepts, year), retrieve, self.log)
            if self.v_arxiv.get() and self.running:
                self._status("searching arXiv", 16, t0)
                found += search_arxiv(concepts, retrieve, year, self.log)
            if self.v_cross.get() and self.running:
                self._status("searching Crossref", 20, t0)
                found += search_crossref(concepts, retrieve, year, self.log)
            if self.v_oa.get() and self.running:
                self._status("searching OpenAlex", 24, t0)
                found += search_openalex(concepts, min(100, retrieve), year, self.log)
            if not self.running:
                self.post("done"); return
            if not found:
                self.log("Nothing found. Drop a term with × to widen the search, "
                         "or clear the year filter.", "err")
                self._status("no results", 0, t0); self.post("done"); return

            papers = dedupe(found, self.log)
            if not self.v_nofl.get():
                before = len(papers)
                papers = [p for p in papers if p["has_fulltext"]]
                self.log(f"Dropped {before - len(papers)} without a full-text link", "dim")
            for i, p in enumerate(papers):
                p["key"] = p.get("doi") or p.get("pmid") or f"k{i}"
                p["scored"] = None

            self._status("ranking locally", 26, t0)
            rank_papers(papers, question, concepts)
            covered = sum(1 for p in papers if (p.get("match") or {}).get("coverage", 0) >= 0.99)
            self.log(f"{covered} of {len(papers)} match every search term", "dim")
            shortlist = papers[:keep]
            self.themes = cluster_themes(shortlist)
            self.papers = papers
            self.post("themes"); self.post("rows")
            self.log(f"{len(papers)} unique records · scoring the top {len(shortlist)}", "ok")

            folder = make_folder(", ".join(concepts))
            self.folder = folder

            if self.v_pdf.get() and self.running:
                self._status("downloading open-access PDFs", 32, t0)
                for i, p in enumerate(shortlist):
                    if not self.running:
                        break
                    p["pdf_path"] = download_pdf(p, folder, self.log)
                    self._status(f"PDFs {i+1}/{len(shortlist)}",
                                 32 + int(10 * (i + 1) / max(1, len(shortlist))), t0)
                got = sum(1 for p in shortlist if p.get("pdf_path"))
                self.log(f"PDFs saved: {got}/{len(shortlist)}", "ok")
                self.post("rows")

            self.ctx = dict(question=question, concepts=concepts, folder=folder,
                            query=pubq, model=model, workers=workers, shortlist=shortlist,
                            profile=profile,
                            sources=", ".join(n for n, on in (
                                ("PubMed", self.v_pubmed.get()), ("Europe PMC", self.v_epmc.get()),
                                ("arXiv", self.v_arxiv.get()), ("Crossref", self.v_cross.get()),
                                ("OpenAlex", self.v_oa.get())) if on))
            self._save_wb()
            secs = int(time.time() - t0)
            self._status(f"{len(papers)} ranked locally in {secs} s — usable now", 100, t0)
            self.log(f"Local ranking finished in {secs} s. Rows are ordered by relevance "
                     f"to your question — no model needed.", "ok")

            mode = self.v_score.get()
            if mode == "off":
                self.log("AI scoring off — local ranking only.", "dim")
            elif mode == "auto":
                self._score_batch(len(shortlist))
            else:
                self.post("ask", n=len(shortlist))
        except Exception as e:
            self.log(f"Unexpected error: {e}", "err")
        finally:
            self.running = False
            self.post("done")

    # ── AI scoring, on demand ─────────────────────────────────────────────────
    def _save_wb(self):
        c = self.ctx
        if not c:
            return
        try:
            path = save_workbook(self.papers, c["folder"], ", ".join(c["concepts"]),
                                 c["question"], c["query"], self.decisions,
                                 c.get("profile", "general"), c.get("sources", ""))
            self.log(f"Workbook saved: {path.name}", "ok")
        except Exception as e:
            self.log(f"Excel error: {e}", "err")

    def _eta_minutes(self, n):
        # measured: ~7 s/paper on a 3b model at 5-per-call with 4 workers
        per_paper = (5.0 if self.v_fast.get() else 9.0)
        batch = max(1, int(self.v_batch.get() or 5))
        workers = max(1, int(self.v_workers.get() or 4))
        gain = 1.0 if batch == 1 else (0.45 if batch >= 5 else 0.6)
        return max(1, int(round(n * per_paper * gain / workers / 60)))

    def _show_ask(self, n):
        """Results are already usable — the model is an explicit, priced choice."""
        T = self.T
        for w in self.ask_bar.winfo_children():
            w.destroy()
        msg = (f"{len(self.papers)} papers ranked locally, ready to read. "
               f"AI scoring adds a 0–12 rubric with quoted evidence — "
               f"about {self._eta_minutes(n)} min for the top {n}.")
        tk.Label(self.ask_bar, text=msg, anchor="w", justify="left", bg=T["surface"],
                 fg=T["text"], font=self.fs(11)).pack(side="left", padx=14, pady=9)
        tk.Button(self.ask_bar, text="NOT NOW", relief="flat", bd=0, cursor="hand2",
                  padx=12, pady=4, bg=T["surface"], fg=T["mut"], highlightthickness=1,
                  highlightbackground=T["line"], activeforeground=T["accent"],
                  font=self.fs(10, kind="mono"),
                  command=self._hide_ask).pack(side="right", padx=(6, 14))
        choices = [(n, True)] if n <= 10 else [(n, True), (10, False)]
        for k, primary in choices:
            tk.Button(self.ask_bar, text=f"SCORE TOP {k}  ~{self._eta_minutes(k)} min",
                      relief="flat", bd=0, cursor="hand2", padx=12, pady=4,
                      bg=T["accent"] if primary else T["surface"],
                      fg=T["on_accent"] if primary else T["text"],
                      highlightthickness=1, highlightbackground=T["line"],
                      activebackground=T["accent_dim"] if primary else T["surface2"],
                      font=self.fs(10, "bold", "head"),
                      command=lambda k=k: self._start_scoring(k)).pack(side="right", padx=3)
        self.ask_bar.pack(fill="x", side="top", before=self.tree)

    def _hide_ask(self):
        self.ask_bar.pack_forget()
        self.log("Skipped AI scoring. Press S on any row to score just that one.", "dim")

    def _start_scoring(self, n):
        self.ask_bar.pack_forget()
        threading.Thread(target=self._score_batch, args=(n,), daemon=True).start()

    def _score_batch(self, n):
        c = self.ctx
        if not c:
            return
        shortlist = c["shortlist"][:n]
        want = self.v_smodel.get().strip() or SCORE_MODEL
        ok, msg = check_ollama(want)
        if not ok:
            # fall back to the drafting model rather than refusing to score
            ok, msg = check_ollama(c["model"])
            if not ok:
                self.log(msg, "err"); self.post("done"); return
            self.log(f"Scoring model not pulled — using {msg}. "
                     f"(ollama pull {want} is ~3× faster.)", "dim")
        model = msg
        fast, workers = self.v_fast.get(), c["workers"]
        batch = max(1, int(self.v_batch.get() or 5))
        groups = [shortlist[i:i + batch] for i in range(0, len(shortlist), batch)]
        self.scoring = self.running = True
        self.post("scoring", on=True)
        self.log(f"Scoring {len(shortlist)} papers · {model} · {batch} per call · "
                 f"{workers} calls at a time{' · fast mode' if fast else ''}", "accent")
        t0, done = time.time(), 0
        with ThreadPoolExecutor(max_workers=workers) as pool:
            futures = {pool.submit(score_group, g, c["question"], model, self.log, fast,
                                   c.get("profile", "general")): g for g in groups}
            for fut in as_completed(futures):
                g = futures[fut]
                if not self.running:
                    break
                try:
                    res = fut.result()
                except Exception as e:
                    res = {}; self.log(f"  scoring failed: {e}", "err")
                for idx, r in (res or {}).items():
                    if 0 <= idx < len(g):
                        g[idx]["scored"] = r
                        self.log(f"  {r['total']}/12  {g[idx]['title'][:66]}", "ok")
                if not res:
                    self.log(f"  unusable answer for {len(g)} papers — retrying singly", "dim")
                    for p in g:
                        if not self.running:
                            break
                        r = score_paper(p, c["question"], model, self.log, fast,
                                        c.get("profile", "general"))
                        if r:
                            p["scored"] = r
                done += len(g)
                per  = (time.time() - t0) / max(1, done)
                left = int(per * (len(shortlist) - done))
                self.post("status",
                          text=f"scored {min(done, len(shortlist))}/{len(shortlist)} · "
                               f"~{left // 60} m {left % 60:02d} s left · {per:.1f} s/paper",
                          pct=int(100 * done / max(1, len(shortlist))),
                          elapsed=f"{int(time.time() - t0)} s")
                self.post("rows"); self.post("detail")
        self.papers.sort(key=lambda x: (x.get("scored") or {}).get("total", -1), reverse=True)
        self.post("rows")
        self._save_wb()
        self.scoring = self.running = False
        self.post("scoring", on=False)
        secs = int(time.time() - t0)
        self._status(f"{sum(1 for p in self.papers if p.get('scored'))} scored in "
                     f"{secs // 60} m {secs % 60:02d} s", 100)
        self.log("Scoring done. Select a row to see why it scored what it did.", "accent")

    def score_one(self):
        """Score just the selected paper — seconds, not minutes."""
        p = self.selected
        if not p or not self.ctx:
            return
        if p.get("scored"):
            self.log("Already scored — rescoring.", "dim")

        def job():
            t0 = time.time()
            res = score_paper(p, self.ctx["question"],
                              self.v_smodel.get().strip() or self.ctx["model"],
                              self.log, self.v_fast.get(),
                              self.ctx.get("profile", "general"))
            if res:
                p["scored"] = res
                self.log(f"  {res['total']}/12 in {time.time()-t0:.0f} s  "
                         f"{p['title'][:60]}", "ok")
                self.post("rows"); self.post("detail")
            else:
                self.log("  model returned nothing usable", "dim")

        self.log(f"Scoring one paper with {self.ctx['model']}…", "dim")
        threading.Thread(target=job, daemon=True).start()

    # ── synthesis ─────────────────────────────────────────────────────────────
    def export_citations(self):
        """RIS + BibTeX of the included papers, or everything if nothing is screened."""
        if not self.papers:
            messagebox.showinfo("Veronica", "Run a search first.")
            return
        picked = [p for p in self.papers if self.decisions.get(p["key"]) == "include"]
        which = "included"
        if not picked:
            picked, which = self.papers, "all retrieved"
        folder = self.folder or make_folder("review")
        stem = sanitize(", ".join((self.ctx or {}).get("concepts") or ["review"]))
        try:
            ris = save_ris(folder / f"Citations_{stem}_{TODAY}.ris", picked)
            bib = save_bibtex(folder / f"Citations_{stem}_{TODAY}.bib", picked)
        except Exception as e:
            self.log(f"Citation export failed: {e}", "err")
            return
        self.log(f"Exported {len(picked)} {which} papers — {ris.name} (Zotero, Mendeley, "
                 f"EndNote) and {bib.name} (LaTeX)", "ok")
        self._open_path(str(folder))

    def open_overview(self):
        """Year and citation distributions, plus a map of the shortlist by theme."""
        if not self.papers:
            messagebox.showinfo("Veronica", "Run a search first.")
            return
        T = self.T
        win = tk.Toplevel(self)
        win.title("Overview")
        W, H = min(1080, int(self.sw * 0.8)), min(760, int(self.sh * 0.8))
        win.geometry(f"{W}x{H}")
        win.configure(bg=T["bg"])
        cv = tk.Canvas(win, bg=T["bg"], highlightthickness=0, bd=0)
        cv.pack(fill="both", expand=True)
        mono = self.fs(9, kind="mono")
        head = self.fs(12, "bold", "head")

        def label(x, y, text, fill=None, font=mono, anchor="w"):
            cv.create_text(x, y, text=text, fill=fill or T["mut"], font=font, anchor=anchor)

        pad = 40
        colw = (W - pad * 3) // 2

        # ── years ──
        years = [int(p["year"]) for p in self.papers
                 if (p.get("year") or "").isdigit() and 1900 < int(p["year"]) < 2100]
        label(pad, 32, "PUBLICATION YEARS", T["accent"], head)
        if years:
            lo, hi = min(years), max(years)
            counts = Counter(years)
            peak = max(counts.values())
            x0, y0, h = pad, 190, 120
            span = max(1, hi - lo)
            barw = max(2, min(22, int(colw / (span + 1)) - 2))
            for yr in range(lo, hi + 1):
                n = counts.get(yr, 0)
                x = x0 + int((yr - lo) / span * (colw - barw)) if span else x0
                bh = int(h * n / peak) if peak else 0
                cv.create_rectangle(x, y0 - bh, x + barw, y0, fill=T["accent"], width=0)
                if n == peak:
                    label(x + barw / 2, y0 - bh - 10, str(n), T["text"], mono, "center")
            cv.create_line(x0, y0, x0 + colw, y0, fill=T["line"])
            label(x0, y0 + 14, str(lo))
            label(x0 + colw, y0 + 14, str(hi), anchor="e")
            label(pad, 56, f"{len(years)} papers · median "
                           f"{sorted(years)[len(years)//2]}", T["mut2"])
        else:
            label(pad, 60, "no usable years", T["mut2"])

        # ── citations ──
        cx = pad * 2 + colw
        label(cx, 32, "CITATION COUNTS", T["accent"], head)
        cites = [p["cited_by"] for p in self.papers if p.get("cited_by") is not None]
        if cites:
            buckets = [("0", 0, 0), ("1-9", 1, 9), ("10-49", 10, 49),
                       ("50-199", 50, 199), ("200+", 200, 10 ** 9)]
            vals = [sum(1 for c in cites if a <= c <= b) for _n, a, b in buckets]
            peak = max(vals) or 1
            y0, h, bw = 190, 120, int(colw / len(buckets)) - 16
            for i, ((name, _a, _b), v) in enumerate(zip(buckets, vals)):
                x = cx + i * (bw + 16)
                bh = int(h * v / peak)
                cv.create_rectangle(x, y0 - bh, x + bw, y0, fill=T["accent_dim"], width=0)
                label(x + bw / 2, y0 - bh - 10, str(v), T["text"], mono, "center")
                label(x + bw / 2, y0 + 14, name, T["mut2"], mono, "center")
            cv.create_line(cx, y0, cx + colw, y0, fill=T["line"])
            label(cx, 56, f"{len(cites)} with citation data · max {max(cites)}", T["mut2"])
        else:
            label(cx, 60, "no citation data — enable OpenAlex or Crossref", T["mut2"])

        # ── evidence map ──
        label(pad, 265, "EVIDENCE MAP", T["accent"], head)
        label(pad, 288, "year across, relevance up, size = score. Click a paper to select it.",
              T["mut2"])
        mx0, my0 = pad + 30, 330
        mw, mh = W - pad * 2 - 40, H - my0 - 70
        cv.create_rectangle(mx0, my0, mx0 + mw, my0 + mh, outline=T["line"])
        pts = [p for p in self.papers if (p.get("year") or "").isdigit()]
        if pts:
            ys = [int(p["year"]) for p in pts]
            ylo, yhi = min(ys), max(ys)
            rmax = max((p.get("prerank") or 0) for p in pts) or 1
            themes = sorted({p.get("theme", "General") for p in pts})
            for p in pts:
                fx = (int(p["year"]) - ylo) / max(1, yhi - ylo)
                fy = (p.get("prerank") or 0) / rmax
                x = mx0 + 12 + fx * (mw - 24)
                y = my0 + mh - 12 - fy * (mh - 24)
                s = p.get("scored") or {}
                r = 4 + (s.get("total", 0) or 0) / 12 * 7
                inc = self.decisions.get(p["key"]) == "include"
                dot = cv.create_oval(x - r, y - r, x + r, y + r,
                                     fill=T["accent"] if inc else T["accent_dim"],
                                     outline=T["text"] if inc else "", width=1)
                tip = f"{p.get('title','')[:70]} ({p.get('year','')})"
                cv.tag_bind(dot, "<Enter>",
                            lambda e, t=tip: (cv.delete("tip"),
                                              cv.create_text(mx0 + 8, my0 - 12, text=t,
                                                             fill=T["text"], font=mono,
                                                             anchor="w", tags="tip")))
                cv.tag_bind(dot, "<Button-1>",
                            lambda e, k=p["key"]: self._select_key(k))
            label(mx0, my0 + mh + 18, f"{ylo}", T["mut2"])
            label(mx0 + mw, my0 + mh + 18, f"{yhi}", T["mut2"], anchor="e")
            label(pad, my0 + mh + 40,
                  f"filled = included · {len(themes)} themes · {len(pts)} papers plotted",
                  T["mut2"])

    def _select_key(self, key):
        try:
            self.tree.selection_set(key)
            self.tree.see(key)
            self.lift()
        except tk.TclError:
            pass

    def open_overview_safe(self):
        try:
            self.open_overview()
        except Exception as e:
            self.log(f"Overview failed: {e}", "err")

    def do_synthesis(self):
        included = [p for p in self.papers
                    if self.decisions.get(p["key"]) == "include"]
        if not included:
            messagebox.showinfo("Veronica", "No papers included yet: select a row and "
                                            "press I (or the INCLUDE button).")
            return
        question = self.txt_q.get("1.0", "end").strip()
        model    = self.v_model.get().strip() or DEFAULT_MODEL
        ok, msg  = check_ollama(model)
        if not ok:
            messagebox.showinfo("Veronica", f"Drafting needs Ollama running.\n\n{msg}")
            return
        model    = msg
        folder   = self.folder or make_folder("review")
        self.log(f"Drafting findings from {len(included)} included papers…", "accent")
        if not any(p.get("scored") for p in included):
            self.log("None are AI-scored — drafting from abstracts instead.", "dim")

        def job():
            text = draft_synthesis(included, question, model, self.log)
            if not text:
                return
            excluded = [p for p in self.papers
                        if self.decisions.get(p["key"]) == "exclude"]
            meta = {"sources": (self.ctx or {}).get("sources", "—"),
                    "query": (self.ctx or {}).get("query", "—"),
                    "retrieved": len(self.papers),
                    "scored": sum(1 for p in included if p.get("scored")),
                    "model": model}
            profile = (self.ctx or {}).get("profile") or self.profile
            stem = sanitize(", ".join((self.ctx or {}).get("concepts") or ["review"]))
            path = folder / f"Review_{stem}_{TODAY}.docx"
            try:
                save_review_docx(path, question, text, included, excluded, profile, meta)
            except Exception as e:
                self.log(f"Could not write the Word file: {e}", "err")
                return
            self.post("synthesis", text=text, path=str(path))

        threading.Thread(target=job, daemon=True).start()

    def _show_synthesis(self, text, path):
        T = self.T
        win = tk.Toplevel(self)
        win.title("Draft review")
        win.geometry(f"{min(900, int(self.sw*0.7))}x{min(760, int(self.sh*0.8))}")
        win.configure(bg=T["bg"])
        bar = tk.Frame(win, bg=T["bg"]); bar.pack(fill="x", padx=16, pady=(14, 8))
        tk.Label(bar, text=Path(path).name, anchor="w", bg=T["bg"], fg=T["text"],
                 font=self.fs(11)).pack(side="left")
        tk.Button(bar, text="OPEN IN WORD", relief="flat", bd=0, cursor="hand2",
                  padx=14, pady=5, bg=T["accent"], fg=T["on_accent"],
                  activebackground=T["accent_dim"], font=self.fs(11, "bold", "head"),
                  command=lambda: self._open_path(path)).pack(side="right")
        tk.Button(bar, text="SHOW FOLDER", relief="flat", bd=0, cursor="hand2",
                  padx=12, pady=5, bg=T["bg"], fg=T["text"], highlightthickness=1,
                  highlightbackground=T["line"], activeforeground=T["accent"],
                  font=self.fs(10, kind="mono"),
                  command=self.open_folder).pack(side="right", padx=8)
        tk.Label(win, text="Preview of the drafted findings. The Word file also carries the "
                           "included-papers table, the evidence quotes behind every score, "
                           "references and a method note.",
                 anchor="w", justify="left", wraplength=int(760 * self.scale),
                 bg=T["bg"], fg=T["mut2"], font=self.fs(10)).pack(fill="x", padx=16)
        box = tk.Text(win, relief="flat", bd=0, wrap="word", bg=T["surface"], fg=T["text"],
                      insertbackground=T["accent"], font=self.fs(12))
        box.pack(fill="both", expand=True, padx=16, pady=16)
        box.insert("end", text)
        self.log(f"Word document saved: {path}", "ok")


if __name__ == "__main__":
    Veronica().mainloop()
